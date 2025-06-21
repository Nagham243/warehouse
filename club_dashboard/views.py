from django.shortcuts import render, redirect,get_object_or_404
from accounts.models import UserProfile, StudentProfile, CoachProfile ,ReceptionistProfile,AdministrativeProfile,AccountantProfile
from django.contrib.auth.models import User
from .forms import StudentProfileForm, ArticleModelForm, ServicesModelForm, ServicesClassificationModelForm, ProductsModelForm, ProductsClassificationModelForm,ReceptionistProfileForm,ProductShipmentForm,AdministratorProfileForm,AccountantProfileForm
from accounts.forms import ReceptionistSignupForm,AdministratorSignupForm
from students.models import Blog, ServicesModel, ServicesClassificationModel, ProductsModel, ProductsClassificationModel, ProductsImage, ServicesImage,Order,OrderItem,ServiceOrderModel,OrderCancellation
from django.utils import timezone
# Create your views here.
from django.contrib import messages  # ✅ Fix missing import
from .forms import DirectorProfileForm  # ✅ Fix missing import
from accounts.models import DirectorProfile  # ✅ Add this import
from .models import Notification  # Import the Notification model
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from accounts.models import UserProfile
from club_dashboard.models import Notification  # ✅ Import Notification Model
from .utils import send_notification  # ✅ Import notification function
from django.contrib.auth.decorators import login_required  # ✅ Fix missing import
from django.db.models import Avg
from club_dashboard.models import Review  # ✅ Import Review model from students app
from .models import SalonAppointment,ProductShipment,DashboardSettings
from django.shortcuts import render
from django.db.models import Sum, F, FloatField, Case, When, IntegerField, Value
from django.db.models.functions import Cast
from .models import ProductsModel ,ProductImg
from django.utils import timezone
from django.contrib import messages
from .forms import ProductsModelForm
from django.core.paginator import Paginator
import base64
import time
import decimal
from django.core.files.base import ContentFile
from django.utils import timezone
from .forms import ServicesModelForm
from datetime import datetime, timedelta
from django.db import models , transaction
from receptionist_dashboard.models import BookingService
from django.template.loader import render_to_string
import json
from django.http import HttpResponseForbidden
from django.urls import reverse
from accounts.models import UserProfile,DirectorProfile
from .forms import DirectorProfileForm
from django.utils import translation
from receptionist_dashboard.models import SalonBooking
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from django.utils.translation import gettext as _
from django.db.models import Count
from accounts.models import ClubsModel

# Helper function to get user's club
def get_user_club(user):
    user_profile = user.userprofile
    club = None
    if user_profile.account_type == '3':  # Student
        club = user_profile.student_profile.club if hasattr(user_profile, 'student_profile') else None
    elif user_profile.account_type == '4':  # Coach
        club = user_profile.Coach_profile.club if hasattr(user_profile, 'Coach_profile') else None
    elif user_profile.account_type == '2':  # Director
        club = user_profile.director_profile.club if hasattr(user_profile, 'director_profile') else None
    elif user_profile.account_type == '5':  # Receptionist
        club = user_profile.receptionist_profile.club if hasattr(user_profile, 'receptionist_profile') else None
    return club

@login_required
def club_dashboard_index(request):
    context = {}
    user = request.user

    # ✅ Ensure the user has a valid director profile
    if not hasattr(user, 'userprofile') or not user.userprofile.director_profile:
        messages.error(request, "Unauthorized access.")
        return redirect('home')

    # ✅ Get the correct club for the director
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)
    club_name = club.name

    club_admin = user.userprofile.director_profile
    # ✅ Get directors linked through UserProfile
    directors = UserProfile.objects.filter(account_type='6', administrator_profile__club=club)
    director_count = directors.count()

    # ✅ Fetch students and coaches from this club
    students = StudentProfile.objects.filter(club=club)

    # Map for Arabic labels
    manual_status_labels = {
        'trial': 'تجريبي',
        'active': 'نشط',
        'expiring_soon': 'سينتهي قريبًا',
        'expired': 'منتهي',
    }

    # Get counts for each manual status
    manual_status_counts = StudentProfile.objects.filter(club=club) \
        .values('manual_status') \
        .annotate(count=Count('manual_status'))

    manual_status_dict = {status: 0 for status in manual_status_labels.keys()}

    for entry in manual_status_counts:
        key = entry['manual_status'] or 'trial'
        if key in manual_status_dict:
            manual_status_dict[key] = entry['count']

    manual_chart_labels = [manual_status_labels[k] for k in manual_status_dict.keys()]
    manual_chart_data = list(manual_status_dict.values())

    student_count = students.count()

    coaches = CoachProfile.objects.filter(club=club)
    coach_count = coaches.count()

    notifications = Notification.objects.filter(club=club, is_read=False).order_by('-created_at')
    unread_count = notifications.count()
    notifications.update(is_read=True)

    active_count, expiring_soon_count, expired_count, trial_count = 0, 0, 0, 0
    for student in students:
        status = student.get_subscription_status()
        if status == "active":
            active_count += 1
        elif status == "expiring_soon":
            expiring_soon_count += 1
        elif status == "trial":
            trial_count += 1
        else:
            expired_count += 1

    total_students = max(1, active_count + expiring_soon_count + expired_count + trial_count)

    def calc_percent(v, total):
        return round((v / total) * 100, 2) if total > 0 else 0

    active_percentage = calc_percent(active_count, total_students)
    expiring_soon_percentage = calc_percent(expiring_soon_count, total_students)
    expired_percentage = calc_percent(expired_count, total_students)
    trial_percentage = calc_percent(trial_count, total_students)

    top_rated_coaches = (
        CoachProfile.objects.filter(club=club)
        .annotate(avg_rating=Avg('coach_reviews__rating'))
        .filter(avg_rating__isnull=False)
        .order_by('-avg_rating')[:5]
    )

    top_reviews = (
        Review.objects.filter(coach__club=club)
        .order_by('-rating', '-created_at')[:5]
    )

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date:
        start_date = (timezone.now() - timezone.timedelta(days=30)).strftime('%Y-%m-%d')
    if not end_date:
        end_date = timezone.now().strftime('%Y-%m-%d')

    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
        end_date_obj = end_date_obj.replace(hour=23, minute=59, second=59)
    except ValueError:
        start_date_obj = timezone.now() - timezone.timedelta(days=30)
        end_date_obj = timezone.now()

    orders = Order.objects.filter(
        club=club,
        created_at__gte=start_date_obj,
        created_at__lte=end_date_obj
    ).order_by('-created_at')


    if club:
        directors = UserProfile.objects.filter(
            account_type='2',
            director_profile__club=club
        ).select_related('user', 'director_profile')

        receptionists = UserProfile.objects.filter(
            account_type='5',
            receptionist_profile__club=club
        ).select_related('user', 'receptionist_profile')

        administrators = UserProfile.objects.filter(
            account_type='6',
            administrator_profile__club=club
        ).select_related('user', 'administrator_profile')

        accountants = UserProfile.objects.filter(
            account_type='7',
            accountant_profile__club=club
        ).select_related('user', 'accountant_profile')

        staff_list = []

        for director in directors:
            staff_list.append({
                'userprofile': director,
                'role': 'مدير عام',
                'role_en': 'General Manager',
                'profile': director.director_profile,
                'profile_type': 'director'
            })

        for receptionist in receptionists:
            staff_list.append({
                'userprofile': receptionist,
                'role': 'موظف استقبال',
                'role_en': 'Receptionist',
                'profile': receptionist.receptionist_profile,
                'profile_type': 'receptionist'
            })

        for administrator in administrators:
            staff_list.append({
                'userprofile': administrator,
                'role': 'إداري',
                'role_en': 'Administrator',
                'profile': administrator.administrator_profile,
                'profile_type': 'administrator'
            })

        for accountant in accountants:
            staff_list.append({
                'userprofile': accountant,
                'role': 'محاسب',
                'role_en': 'Accountant',
                'profile': accountant.accountant_profile,
                'profile_type': 'accountant'
            })

        staff_list.sort(key=lambda x: x['userprofile'].creation_date, reverse=True)

    else:
        staff_list = []



    total_revenue = int(orders.filter(status__in=['confirmed', 'completed']).aggregate(Sum('total_price'))['total_price__sum'] or 0)
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/index.html', {
        'staff_list':staff_list,
        'clubName': club_name,
        'students': students,
        'coaches': coaches,
        'directors': directors,
        'director_count': director_count,
        'coach_count': coach_count,
        'student_count': student_count,
        'active_count': active_count,
        'expiring_soon_count': expiring_soon_count,
        'expired_count': expired_count,
        'trial_count': trial_count,
        'active_percentage': active_percentage,
        'expiring_soon_percentage': expiring_soon_percentage,
        'expired_percentage': expired_percentage,
        'trial_percentage': trial_percentage,
        'notifications': notifications,
        'unread_count': unread_count,
        'top_rated_coaches': top_rated_coaches,
        'top_reviews': top_reviews,
        'total_revenue': total_revenue,
        'club_admin': club_admin,
        'manual_chart_labels': manual_chart_labels,
        'manual_chart_data': manual_chart_data,
    })

@login_required
def delete_review(request, review_id):
    review = get_object_or_404(Review, id=review_id)

    user = request.user
    if hasattr(user, 'userprofile') and user.userprofile.director_profile:
        if review.coach.club == user.userprofile.director_profile.club:
            review.delete()
            messages.success(request, "Review deleted successfully.")
        else:
            messages.error(request, "You are not authorized to delete this review.")
    else:
        messages.error(request, "Unauthorized action.")

    return redirect('club_dashboard_index')  # Use your correct dashboard URL name

def viewStudents(request):
    context ={}
    """Displays all students in the club."""
    user = request.user

    # if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
    #     messages.error(request, "Unauthorized access.")
    #     return redirect('home')

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)


    students = UserProfile.objects.filter(
        account_type='3', 
        student_profile__club=club
    ).select_related('user', 'student_profile')

    # ✅ Enrich each student with subscription status
    for student in students:
        profile = student.student_profile
        if profile and hasattr(profile, 'get_subscription_status'):
            student.subscription_status = profile.get_subscription_status()
            student.manual_status_display = profile.get_manual_status_display()
        else:
            student.subscription_status = "unknown"
            student.manual_status_display = "-"
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/students/viewStudents.html', {'students': students,'club': club})



import pandas as pd
import openpyxl
from django.contrib import messages
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.models import User
from django.db import transaction
from datetime import datetime
import io

# Add this view for handling the import
def import_students(request):
    """Display import students page"""
    context = {}
    user = request.user

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    if not club:
        messages.error(request, "No club associated with your account.")
        return redirect('viewStudents')

    context['club'] = club
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/students/import_students.html', context)



def process_import_students(request):
    """Process the uploaded Excel/CSV file"""
    if request.method != 'POST':
        return redirect('import_students')

    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    if not club:
        messages.error(request, "No club associated with your account.")
        return redirect('viewStudents')

    if 'file' not in request.FILES:
        messages.error(request, "No file uploaded.")
        return redirect('import_students')

    file = request.FILES['file']

    # Validate file type
    if not file.name.endswith(('.xlsx', '.xls', '.csv')):
        messages.error(request, "Please upload an Excel (.xlsx, .xls) or CSV file.")
        return redirect('import_students')

    try:
        # Read the file
        if file.name.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)

        # Validate required columns
        required_columns = ['username', 'email', 'full_name', 'phone', 'birthday']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            messages.error(request, f"Missing required columns: {', '.join(missing_columns)}")
            return redirect('import_students')

        # Process the data
        success_count = 0
        error_count = 0
        errors = []

        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    # Check if username already exists
                    if User.objects.filter(username=row['username']).exists():
                        errors.append(f"Row {index + 2}: Username '{row['username']}' already exists")
                        error_count += 1
                        continue

                    # Check if email already exists
                    if User.objects.filter(email=row['email']).exists():
                        errors.append(f"Row {index + 2}: Email '{row['email']}' already exists")
                        error_count += 1
                        continue

                    # Parse birthday
                    try:
                        birthday = pd.to_datetime(row['birthday']).date()
                    except:
                        errors.append(f"Row {index + 2}: Invalid birthday format")
                        error_count += 1
                        continue

                    # Create User
                    user_obj = User.objects.create_user(
                        username=row['username'],
                        email=row['email'],
                        password='defaultpass123'  # You might want to generate random passwords
                    )

                    # Create StudentProfile
                    student_profile = StudentProfile.objects.create(
                        full_name=row['full_name'],
                        phone=str(row['phone']),
                        birthday=birthday,
                        club=club,
                        manual_status=row.get('manual_status', 'trial')
                    )

                    # Create UserProfile
                    UserProfile.objects.create(
                        user=user_obj,
                        account_type='3',  # student
                        student_profile=student_profile,
                        is_active=True
                    )

                    success_count += 1

                except Exception as e:
                    errors.append(f"Row {index + 2}: {str(e)}")
                    error_count += 1

        # Show results
        if success_count > 0:
            messages.success(request, f"Successfully imported {success_count} students.")

        if error_count > 0:
            error_message = f"Failed to import {error_count} students:\n" + "\n".join(errors[:10])
            if len(errors) > 10:
                error_message += f"\n... and {len(errors) - 10} more errors"
            messages.error(request, error_message)

    except Exception as e:
        messages.error(request, f"Error processing file: {str(e)}")

    return redirect('viewStudents')

def download_sample_template(request):
    """Download a sample Excel template for importing students"""

    # Create sample data
    data = {
        'username': ['student1', 'student2', 'student3'],
        'email': ['student1@example.com', 'student2@example.com', 'student3@example.com'],
        'full_name': ['Ahmed Ali', 'Sara Mohamed', 'Omar Hassan'],
        'phone': ['01234567890', '01098765432', '01156789012'],
        'birthday': ['1995-01-15', '1998-03-22', '2000-07-10'],
        'manual_status': ['trial', 'active', 'trial']
    }

    df = pd.DataFrame(data)

    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Students', index=False)

        # Format the worksheet
        worksheet = writer.sheets['Students']
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="students_import_template.xlsx"'

    return response




def export_students_excel(request):
    user = request.user

    # if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
    #     messages.error(request, "Unauthorized access.")
    #     return redirect('home')

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    students = UserProfile.objects.filter(
        account_type='3',
        student_profile__club=club
    ).select_related('user', 'student_profile')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "اللاعبين"

    arabic_font = Font(name='Arial', size=12)
    right_align = Alignment(horizontal='right')

    headers = [
        _('Username'), _('Email'), _('Full Name'),
        _('Phone'), _('Birthday'), _('Subscription Status')
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = arabic_font
        cell.alignment = right_align

    for student in students:
        profile = student.student_profile
        subscription = profile.get_subscription_status() if hasattr(profile, 'get_subscription_status') else "unknown"

        row = [
            student.user.username,
            student.user.email,
            profile.full_name if profile else '',
            profile.phone if profile else '',
            str(profile.birthday) if profile and profile.birthday else '',
            subscription
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = arabic_font
            cell.alignment = right_align

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=players.xlsx'
    wb.save(response)
    return response




def addStudent(request):
    context ={}
    """Adds a new student to the club."""
    user = request.user

    # ✅ Ensure user is a director
    # if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
    #     messages.error(request, "Unauthorized access.")
    #     return redirect('home')

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    if not club.can_add_more_players:
        messages.error(request, f"لا يمكن إضافة المزيد من العملاء. الحد الأقصى للباقة الحالية هو {club.get_max_players()} عميل.")
        notification_msg = (
            f"تم الوصول إلى الحد الأقصى للعملاء ({club.get_max_players()}) في المنصة. "
            f"يرجى ترقية الباقة لإضافة المزيد من العملاء."
        )
        send_notification(user, club, notification_msg)

        return redirect('viewStudents')

    form = StudentProfileForm()

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')

        # ✅ Prevent duplicate username and email
        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return redirect('addStudent')

        if User.objects.filter(email=email).exists():
            messages.error(request, "Email is already in use.")
            return redirect('addStudent')

        form = StudentProfileForm(request.POST)
        if form.is_valid():
            # ✅ Create new user
            student = User.objects.create(username=username, email=email)
            if password:
                student.set_password(password)
            student.save()

            # ✅ Create and link Student Profile
            student_profile = form.save(commit=False)
            student_profile.user = student
            student_profile.club = club
            student_profile.save()

            # ✅ Create UserProfile entry
            UserProfile.objects.create(user=student, account_type='3', student_profile=student_profile)

            # ✅ Send notification
            send_notification(user, club, f" العميل الجديد 📢 {username} انضم إلى {club.name}.")

            messages.success(request, "CLient added successfully.")
            return redirect('viewStudents')
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/students/addStudent.html', {'form': form,'club': club})


@login_required
def editStudent(request, id):
    context = {}
    """Edits an existing student's details."""
    user = request.user

    # ✅ Ensure user is a director
    # if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
    #     messages.error(request, "Unauthorized access.")
    #     return redirect('home')

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    student_profile = get_object_or_404(StudentProfile, id=id)
    student = get_object_or_404(User, userprofile__student_profile=student_profile)

    form = StudentProfileForm(instance=student_profile)

    if request.method == 'POST':
        new_username = request.POST.get('username')
        new_email = request.POST.get('email')
        password = request.POST.get('password')

        form = StudentProfileForm(request.POST, instance=student_profile)
        if form.is_valid():
            # ✅ Update student details
            student.username = new_username
            student.email = new_email
            if password:
                student.set_password(password)
            student.save()

            student_profile = form.save(commit=False)
            student_profile.user = student  # 🔥 FIXED: Ensure user link remains
            student_profile.save()

            # ✅ Send notification
            send_notification(user, club, f" تم تحديث ملف العميل 📝 {student.username}.")

            messages.success(request, "CLient profile updated successfully.")
            return redirect('viewStudents')
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/students/editStudent.html', {
        'form': form,
        'student': student,
        'club': club,
    })



@login_required
def deleteStudent(request, id):
    """Deletes a student from the club."""
    user = request.user

    # ✅ Ensure user is a director
    # if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
    #     messages.error(request, "Unauthorized access.")
    #     return redirect('home')

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    student_profile = get_object_or_404(StudentProfile, id=id)
    student = get_object_or_404(User, userprofile__student_profile=student_profile)

    student_name = student.username

    # ✅ Delete student profile and user account
    student_profile.delete()
    student.delete()

    # ✅ Send notification
    send_notification(user, club, f" تم حذف العميل 🗑️ {student_name} .من المنصة ")

    messages.success(request, "CLient has been deleted successfully.")
    return redirect('viewStudents')


from django.db.models import Sum, Q, OuterRef, Subquery, DecimalField
from django.db.models.functions import Coalesce
def viewCoachs(request):
    context = {}
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    # Approved coaches
    coach_userprofile = UserProfile.objects.filter(
        account_type='4',
        Coach_profile__club=club
    ).select_related('user', 'Coach_profile').annotate(
        total_sales=Coalesce(
            Subquery(
                OrderItem.objects.filter(
                    Q(order__status='confirmed'),
                    Q(product__creator__userprofile__Coach_profile=OuterRef('Coach_profile')) |
                    Q(service__creator__userprofile__Coach_profile=OuterRef('Coach_profile'))
                ).values('product__creator__userprofile__Coach_profile')
                .annotate(total=Sum('price'))
                .values('total')[:1]
            ),
            0.00,
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )
    )

    # Pending coaches (for the top section)
    pending_coaches = CoachProfile.objects.filter(
        club=club,
        approval_status='pending'
    ).order_by('-created_at')[:3]  # Get only 3 most recent pending coaches

    # Statistics
    total_coaches = coach_userprofile.count()
    active_coaches = coach_userprofile.filter(is_active=True).count()
    inactive_coaches = total_coaches - active_coaches
    pending_count = CoachProfile.objects.filter(club=club, approval_status='pending').count()

    context.update({
        'LANGUAGE_CODE': translation.get_language(),
        'coach_userprofile': coach_userprofile,
        'club': club,
        'pending_coaches': pending_coaches,
        'stats': {
            'total': total_coaches,
            'active': active_coaches,
            'inactive': inactive_coaches,
            'pending': pending_count,
        }
    })

    return render(request, 'club_dashboard/coachs/viewCoachs.html', context)


def export_coaches_excel(request):
    user = request.user

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    coaches = UserProfile.objects.filter(
        account_type='4',
        Coach_profile__club=club
    ).select_related('user', 'Coach_profile')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _("Coaches")

    font = Font(name='Arial', size=12)
    align = Alignment(horizontal='right')

    headers = [
        _('Username'),
        _('Email'),
        _('Full Name'),
        _('Phone'),
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = font
        cell.alignment = align

    for coach in coaches:
        profile = coach.Coach_profile
        row = [
            coach.user.username,
            coach.user.email,
            profile.full_name if profile else '',
            profile.phone if profile else '',
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = font
            cell.alignment = align

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=coaches.xlsx'
    wb.save(response)
    return response

from club_dashboard.forms import CoachProfileForm
@login_required
def addCoach(request):
    context = {}
    user = request.user

    # Get the club associated with the user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    form = CoachProfileForm()

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')

        # Check for duplicate username and email
        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return redirect('addCoach')

        if User.objects.filter(email=email).exists():
            messages.error(request, "Email is already in use.")
            return redirect('addCoach')

        form = CoachProfileForm(request.POST)
        if form.is_valid():
            # Create new user
            coach = User.objects.create(username=username, email=email)
            if password:
                coach.set_password(password)
            coach.save()

            # Create coach profile with approved status
            coach_profile = form.save(commit=False)
            coach_profile.club = club
            coach_profile.approval_status = 'approved'
            coach_profile.approved_by = user
            coach_profile.approved_at = timezone.now()
            coach_profile.save()

            # Create UserProfile entry
            UserProfile.objects.create(
                user=coach,
                account_type='4',
                Coach_profile=coach_profile,
                is_active=True
            )

            # Send notification
            send_notification(user, club, f" التاجر الجديد 📢 {username} انضم إلى {club.name}.")

            messages.success(request, "تم إضافة التاجر بنجاح.")
            return redirect('viewCoachs')

    context['LANGUAGE_CODE'] = translation.get_language()
    context['form'] = form
    context['club'] = club
    return render(request, 'club_dashboard/coachs/addCoach.html', context)


# views.py
@login_required
def editCoach(request, id):
    context = {}
    """Edits an existing coach's details."""
    user = request.user

    # ✅ Ensure user is a director
    # if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
    #     messages.error(request, "Unauthorized access.")
    #     return redirect('home')

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    coach_profile = get_object_or_404(CoachProfile, id=id)
    coach = get_object_or_404(User, userprofile__Coach_profile=coach_profile)

    form = CoachProfileForm(instance=coach_profile, club=club)

    if request.method == 'POST':
        new_username = request.POST.get('username')
        new_email = request.POST.get('email')
        password = request.POST.get('password')

        form = CoachProfileForm(request.POST, instance=coach_profile, club=club)
        if form.is_valid():
            # ✅ Check if username or email was changed
            username_changed = new_username != coach.username
            email_changed = new_email != coach.email

            # Check if vendor classification changed
            old_classification = coach_profile.vendor_classification
            new_classification = form.cleaned_data.get('vendor_classification')
            classification_changed = old_classification != new_classification

            # ✅ Update coach details
            coach.username = new_username
            coach.email = new_email
            if password:
                coach.set_password(password)
            coach.save()

            # Save the form (this will update vendor_classification)
            updated_profile = form.save()

            # Update commission assignment if classification changed
            if classification_changed:
                try:
                    # Get the new commission for this classification
                    new_commission = Commission.objects.filter(
                        club=club,
                        commission_type='vendor',
                        vendor_classification=new_classification,
                        is_active=True
                    ).first()

                    if new_commission:
                        # Update or create commission assignment
                        from club_dashboard.models import VendorCommissionAssignment
                        assignment, created = VendorCommissionAssignment.objects.get_or_create(
                            vendor=coach_profile,
                            defaults={'commission': new_commission}
                        )

                        if not created:
                            assignment.commission = new_commission
                            assignment.save()

                        messages.success(request, f"تم تحديث تصنيف البائع إلى {new_classification} وتم تحديث العمولة.")
                    else:
                        messages.warning(request, f"تم تحديث التصنيف ولكن لم يتم العثور على عمولة مناسبة للتصنيف {new_classification}")

                except Exception as e:
                    messages.error(request, f"حدث خطأ أثناء تحديث العمولة: {str(e)}")

            # ✅ Send notification if changes were made
            notification_message = f" تم تحديث ملف التاجر 📝 {coach.username}."
            if username_changed or email_changed:
                notification_message += " (Username/Email updated.)"
            if classification_changed:
                notification_message += f" (Classification changed from {old_classification} to {new_classification})"

            send_notification(user, club, notification_message)

            if not classification_changed:
                messages.success(request, "Employee profile updated successfully.")
            return redirect('viewCoachs')
        else:
            messages.error(request, "يرجى تصحيح الأخطاء في النموذج.")

    context['LANGUAGE_CODE'] = translation.get_language()
    context.update({
        'form': form,
        'coach': coach,
        'coach_profile': coach_profile,
        'club': club,
    })

    return render(request, 'club_dashboard/coachs/editCoach.html', context)

@login_required
def deleteCoach(request, id):
    """Deletes a coach from the club."""
    user = request.user

    # ✅ Ensure user is a director
    # if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
    #     messages.error(request, "Unauthorized access.")
    #     return redirect('home')

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    coach_profile = get_object_or_404(CoachProfile, id=id)
    coach = get_object_or_404(User, userprofile__Coach_profile=coach_profile)

    coach_name = coach.username

    # ✅ Delete coach profile and user account
    coach_profile.delete()
    coach.delete()

    # ✅ Send notification
    send_notification(user, club, f" تم حذف التاجر 🗑️ {coach_name} .من المنصة ")

    messages.success(request, "Employee has been deleted successfully.")
    return redirect('viewCoachs')



    
def viewAdmins(request):
    user = request.user
    admin_userprofile = UserProfile.objects.filter(account_type='3', Admin_profile__club=user.userprofile.director_profile.club)
    return render(request, 'club_dashboard/admins/viewAdmins.html', {'admin_userprofile': admin_userprofile})
def addAdmin(request):
    user = request.user
    club = user.userprofile.director_profile.club

    form = AdminProfileForm()
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')

        form = AdminProfileForm(request.POST)
        if form.is_valid():
            admin = User.objects.create(username=username, email=email)
            if password:
                admin.set_password(password)
            admin.save()

            admin_profile = form.save()
            admin_profile.club = club
            admin_profile.save()

            userprofile = UserProfile.objects.create(user=admin, account_type='3', Admin_profile=admin_profile)
            userprofile.save()

            return redirect('viewAdmins')

    return render(request, 'club_dashboard/admins/addAdmin.html', {'form': form})

def editAdmin(request, id):
    admin_profile = AdminProfile.objects.get(id=id)
    form = AdminProfileForm(instance=admin_profile)
    admin = User.objects.get(userprofile__Admin_profile=admin_profile)
    username = admin.username
    email = admin.email

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')

        form = AdminProfileForm(request.POST, instance=admin_profile)
        if form.is_valid():
            admin.username = username
            admin.email = email
            if password:
                admin.set_password(password)
            admin.save()

            admin_profile = form.save()

            return redirect('viewAdmins')

    return render(request, 'club_dashboard/admins/editAdmin.html', {'admin': admin, 'form': form, 'email': email, 'username': username})

def deleteAdmin(request, id):
    admin_profile = AdminProfile.objects.get(id=id)
    admin = User.objects.get(userprofile__Admin_profile=admin_profile)

    admin_profile.delete()
    admin.delete()

    return redirect('viewAdmins')




def addProduct(request):
    context = {}
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)
    form = ProductsModelForm()

    if request.method == 'POST':
        form = ProductsModelForm(data=request.POST)
        if form.is_valid():
            ser = form.save(commit=False)
            ser.club = club
            ser.creator = user
            ser.creation_date = timezone.now()
            ser.save()

            profile_imgs = request.POST.getlist('profile_imgs')
            for img_data in profile_imgs:
                format, imgstr = img_data.split(';base64,') if ';base64,' in img_data else ('', img_data)
                ext = format.split('/')[-1] if '/' in format else 'png'

                from django.core.files.base import ContentFile
                import base64, uuid

                data = ContentFile(base64.b64decode(imgstr), name=f'{uuid.uuid4()}.{ext}')

                ProductImg.objects.create(
                    product=ser,
                    img=data
                )

            messages.success(request, 'تم إضافة المنتج بنجاح!')
            return redirect('viewProducts')
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/products/ProductsStock/addProductStock.html', {'form': form,'club': club})

def editProduct(request, id):
    context = {}
    user = request.user
    product = ProductsModel.objects.get(id=id)
    profile_img_objs = ProductImg.objects.filter(product=product)
    form = ProductsModelForm(instance=product)
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    if request.method == 'POST':
        form = ProductsModelForm(data=request.POST, instance=product)
        if form.is_valid():
            profile_imgs = request.POST.getlist('profile_imgs')
            images_changed = request.POST.get('images_changed', 'false') == 'true'

            updated_product = form.save(commit=False)
            updated_product.updated_at = timezone.now()
            updated_product.save()
            form.save_m2m()

            if images_changed:
                ProductImg.objects.filter(product=product).delete()

                for img_data in profile_imgs:
                    if img_data.startswith('data:image'):
                        format, imgstr = img_data.split(';base64,')
                        ext = format.split('/')[-1]

                        import uuid
                        filename = f"{uuid.uuid4()}.{ext}"

                        from django.core.files.base import ContentFile
                        import base64
                        data = ContentFile(base64.b64decode(imgstr))

                        img_obj = ProductImg(product=product)
                        img_obj.img.save(filename, data, save=False)
                        img_obj.save()

            messages.success(request, 'تم تعديل المنتج بنجاح')
            return redirect('viewProducts')

    context = {
        'form': form,
        'profile_imgs': profile_img_objs,
        'club' :club,
    }
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/products/ProductsStock/editProductStock.html', context)

def viewProducts(request):
    context = {}
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)
    products = ProductsModel.objects.filter(club=club)
    total_products = products.count()

    total_value = 0
    low_stock_count = 0
    out_of_stock_count = 0
    expiring_soon_count = 0
    expired_count = 0
    low_stock_threshold = 10

    for product in products:
        product_value = product.price * product.stock
        total_value += product_value

        if 0 < product.stock <= low_stock_threshold:
            low_stock_count += 1

        if product.stock == 0:
            out_of_stock_count += 1

        if product.is_expiring_soon:
            expiring_soon_count += 1

        if product.is_expired:
            expired_count += 1

    paginator = Paginator(products, 6)
    page_number = request.GET.get('page', 1)
    paginated_products = paginator.get_page(page_number)

    context = {
        'products': paginated_products,
        'total_products': total_products,
        'total_value': total_value,
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,
        'expiring_soon_count': expiring_soon_count,
        'expired_count': expired_count,
        'low_stock_threshold': low_stock_threshold,
        'club':club
    }
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/products/ProductsStock/viewProductsStock.html', context)


def DeleteProduct(request, id):
    art = get_object_or_404(ProductsModel, id=id)
    art.delete()
    messages.success(request, 'تم حذف المنتج بنجاح!')
    return redirect('viewProducts')

def addProductClassification(request):
    user = request.user
    club = user.userprofile.director_profile.club
    form = ProductsClassificationModelForm()
    if request.method == 'POST':
        form = ProductsClassificationModelForm(data=request.POST)
        if form.is_valid():
            cla = form.save(commit=False)
            cla.club = club
            cla.creator = user
            cla.creation_date = timezone.now()
            cla.save()

    return render(request, 'club_dashboard/products/Classification/addClassification.html', {'form':form})

def editProductClassification(request, id):
    cla = ProductsClassificationModel.objects.get(id=id)
    form = ProductsClassificationModelForm(instance=cla)
    if request.method == 'POST':
        form = ProductsClassificationModelForm(data=request.POST, instance=cla)
        if form.is_valid():
            form.save()
    return render(request, 'club_dashboard/products/Classification/editClassification.html', {'form':form})

def viewProductsClassification(request):
    classifications = ProductsClassificationModel.objects.all()

    return render(request, 'club_dashboard/products/Classification/viewClassification.html', {'classifications':classifications})

def DeleteProductsClassification(request, id):
    art = ProductsClassificationModel.objects.get(id=id)
    art.delete()
    return redirect('viewProductsClassification')


def addServices(request):
    context = {}
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    coaches = CoachProfile.objects.filter(club=club)
    classifications = ServicesClassificationModel.objects.filter(club=club)

    form = ServicesModelForm()
    form.fields['coaches'].queryset = coaches
    form.fields['classification'].queryset = classifications

    if request.method == 'POST':
        form = ServicesModelForm(data=request.POST)
        form.fields['coaches'].queryset = coaches
        form.fields['classification'].queryset = classifications

        if form.is_valid():
            ser = form.save(commit=False)
            ser.club = club
            ser.creator = user
            ser.creation_date = timezone.now()

            ser.age_from = 0
            ser.age_to = 100
            ser.subscription_days = 30

            duration = request.POST.get('duration')
            if duration:
                ser.duration = int(duration)
            else:
                ser.duration = 0

            discounted_price = request.POST.get('discounted_price')
            if discounted_price and discounted_price.strip():
                ser.discounted_price = decimal.Decimal(discounted_price)

            ser.save()

            # Handle coaches (many-to-many)
            form.save_m2m()

            # Handle classification (single selection for many-to-many field)
            selected_classification = form.cleaned_data.get('classification')
            if selected_classification:
                ser.classification.set([selected_classification])

            image_data = request.POST.get('service_image_data')
            if image_data and image_data.startswith('data:image'):
                format, imgstr = image_data.split(';base64,')
                ext = format.split('/')[-1]

                filename = f"service_{ser.id}_{int(time.time())}.{ext}"
                temp_file = ContentFile(base64.b64decode(imgstr), name=filename)

                ser.image.save(filename, temp_file, save=True)

            return redirect('viewServices')
        else:
            print(form.errors)

    context['LANGUAGE_CODE'] = translation.get_language()
    selected_coach_ids = request.POST.getlist('coaches') if request.method == 'POST' else []
    selected_classification_id = request.POST.get('classification') if request.method == 'POST' else None

    return render(request, 'club_dashboard/services/addServices.html', {
        'form': form,
        'selected_coach_ids': selected_coach_ids,
        'selected_classification_id': selected_classification_id,
        'club': club
    })



def editServices(request, id):
    context = {}
    ser = ServicesModel.objects.get(id=id)
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)
    classifications = ServicesClassificationModel.objects.filter(club=club)

    coaches = CoachProfile.objects.filter(club=club)
    form = ServicesModelForm(instance=ser)
    form.fields['coaches'].queryset = coaches
    form.fields['classification'].queryset = classifications

    if request.method == 'POST':
        form = ServicesModelForm(data=request.POST, instance=ser)
        form.fields['coaches'].queryset = coaches
        form.fields['classification'].queryset = classifications
        if form.is_valid():
            ser = form.save(commit=False)
            ser.creation_date = timezone.now()

            duration = request.POST.get('duration')
            ser.duration = int(duration) if duration else 0

            discounted_price = request.POST.get('discounted_price')
            if discounted_price and discounted_price.strip():
                ser.discounted_price = decimal.Decimal(discounted_price)
            else:
                ser.discounted_price = None

            # Check if the current image should be removed
            remove_current_image = request.POST.get('remove_current_image')
            if remove_current_image == 'true' and ser.image:
                # Delete the old image file
                ser.image.delete(save=False)

            # Handle classification (single selection for many-to-many field)
            selected_classification = form.cleaned_data.get('classification')
            if selected_classification:
                ser.classification.set([selected_classification])

            # Process new image upload if available
            image_data = request.POST.get('service_image_data')
            if image_data and image_data.startswith('data:image'):
                # Get the format and the actual base64 data
                format, imgstr = image_data.split(';base64,')
                ext = format.split('/')[-1]

                # Generate filename and save path
                filename = f"service_{ser.id}_{int(time.time())}.{ext}"
                temp_file = ContentFile(base64.b64decode(imgstr), name=filename)

                # If there's an existing image, delete it first
                if ser.image:
                    ser.image.delete(save=False)

                # Save to the model's ImageField
                ser.image.save(filename, temp_file, save=False)

            ser.save()
            form.save_m2m()
            return redirect('viewServices')
        else:
            print(form.errors)

    context['LANGUAGE_CODE'] = translation.get_language()

    # Get the current coaches and classification for the service
    current_coaches = ser.coaches.all()
    selected_coach_ids = [str(coach.id) for coach in current_coaches]

    # Get the current classification (assuming single selection)
    current_classification = ser.classification.first()
    selected_classification_id = str(current_classification.id) if current_classification else None

    # Add pricing period choices to context
    context.update({
        'form': form,
        'selected_coach_ids': selected_coach_ids,
        'club': club,
        'selected_classification_id': selected_classification_id,
        'pricing_period_choices': ServicesModel.PRICING_PERIOD_CHOICES,
    })

    return render(request, 'club_dashboard/services/editServices.html', context)


def viewServices(request):
    context = {}
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)
    services = ServicesModel.objects.filter(
        club=getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)
    )

    if services:
        # Calculate average monthly price (normalize all prices to monthly rate)
        total_monthly_price = sum(service.monthly_price for service in services)
        avg_monthly_price = total_monthly_price / len(services)
        avg_monthly_price = round(avg_monthly_price, 1)

        # Calculate average duration
        avg_duration = sum(service.duration for service in services) / len(services)
        avg_duration_hours = int(avg_duration // 60)
        avg_duration_minutes = int(avg_duration % 60)

        # Calculate pricing period statistics
        pricing_periods = [service.pricing_period_months for service in services]
        most_common_period = max(set(pricing_periods), key=pricing_periods.count)

        # Get pricing period choices for display
        pricing_period_choices = dict(ServicesModel.PRICING_PERIOD_CHOICES)

    else:
        avg_monthly_price = 0
        avg_duration_hours = 0
        avg_duration_minutes = 0
        most_common_period = 1
        pricing_period_choices = dict(ServicesModel.PRICING_PERIOD_CHOICES)

    context = {
        'services': services,
        'avg_monthly_price': avg_monthly_price,
        'avg_duration_hours': avg_duration_hours,
        'avg_duration_minutes': avg_duration_minutes,
        'most_common_period': most_common_period,
        'pricing_period_choices': pricing_period_choices,
        'club': club,
    }
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/services/viewServices.html', context)

def viewServiceDetails(request, service_id):
    context = {}
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    try:
        service = ServicesModel.objects.get(id=service_id, club=club)
    except ServicesModel.DoesNotExist:
        messages.error(request, 'Service not found.')
        return redirect('viewServices')

    context = {
        'service': service,
        'club': club,
    }
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/services/viewServiceDetails.html', context)

def DeleteServices(request, id):
    art = ServicesModel.objects.get(id=id)
    art.delete()
    return redirect('viewServices')

def addServicesClassification(request):
    user = request.user
    club = user.userprofile.director_profile.club
    form = ServicesClassificationModelForm()
    if request.method == 'POST':
        form = ServicesClassificationModelForm(data=request.POST)
        if form.is_valid():
            cla = form.save(commit=False)
            cla.club = club
            cla.creator = user
            cla.creation_date = timezone.now()
            cla.save()


    return render(request, 'club_dashboard/services/Classification/addClassification.html', {'form':form})

def editServicesClassification(request, id):
    cla = ServicesClassificationModel.objects.get(id=id)
    form = ServicesClassificationModelForm(instance=cla)
    if request.method == 'POST':
        form = ServicesClassificationModelForm(data=request.POST, instance=cla)
        if form.is_valid():
            form.save()

    return render(request, 'club_dashboard/services/Classification/editClassification.html', {'form':form})

def viewServicesClassification(request):
    user = request.user
    club = user.userprofile.director_profile.club
    classifications = ServicesClassificationModel.objects.filter(club=club)
    return render(request, 'club_dashboard/services/Classification/viewClassification.html', {'classifications':classifications})

def DeleteServicesClassification(request, id):
    art = ServicesClassificationModel.objects.get(id=id)
    art.delete()
    return redirect('viewServicesClassification')

#Blog
from django.shortcuts import render, redirect
from django.utils import timezone, translation
from .forms import ArticleModelForm

def addArticle(request):
    context = {}
    user = request.user
    club = user.userprofile.director_profile.club
    form = ArticleModelForm()

    if request.method == 'POST':
        form = ArticleModelForm(data=request.POST, files=request.FILES)
        if form.is_valid():
            art = form.save(commit=False)
            art.club = club
            art.creator = user
            art.creation_date = timezone.now()
            art.save()
            return redirect('viewArticles')

    context['form'] = form
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/blog/addArticle.html', context)

# views.py - Updated editArticle view
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import translation
from django.contrib import messages

def editArticle(request, id):
    user = request.user
    club = user.userprofile.director_profile.club

    # Use get_object_or_404 for better error handling
    art = get_object_or_404(Blog, id=id, club=club)

    if request.method == 'POST':
        form = ArticleModelForm(data=request.POST, files=request.FILES, instance=art)
        if form.is_valid():
            updated_article = form.save(commit=False)
            # Ensure the club and creator remain the same
            updated_article.club = club
            updated_article.creator = art.creator  # Keep original creator
            updated_article.save()

            # Add success message
            if translation.get_language() == 'ar':
                messages.success(request, 'تم تحديث المقال بنجاح!')
            else:
                messages.success(request, 'Article updated successfully!')

            return redirect('viewArticles')
        else:
            # Add error message if form is invalid
            if translation.get_language() == 'ar':
                messages.error(request, 'يرجى تصحيح الأخطاء أدناه.')
            else:
                messages.error(request, 'Please correct the errors below.')
    else:
        form = ArticleModelForm(instance=art)

    context = {
        'form': form,
        'article': art,  # Pass the article instance for additional template usage
        'LANGUAGE_CODE': translation.get_language(),
    }

    return render(request, 'club_dashboard/blog/editArticle.html', context)

def viewArticles(request):
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    arts = Blog.objects.filter(club=club)  # Only show articles from this club

    # Get statistics for the dashboard
    total_articles = arts.count()
    current_month = timezone.now().month
    current_year = timezone.now().year
    new_articles_this_month = arts.filter(
        creation_date__month=current_month,
        creation_date__year=current_year
    ).count()

    # Get most popular articles (assuming you have a views or likes field)
    # If not, you can add this feature later
    popular_articles = arts.order_by('-id')[:3]

    context = {
        'arts': arts,
        'total_articles': total_articles,
        'new_articles_this_month': new_articles_this_month,
        'popular_articles': popular_articles.count(),
        'club':club,
    }
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/blog/viewArticless.html', context)

def DeleteArticle(request, id):
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    try:
        art = Blog.objects.get(id=id, club=club)
        art.delete()
    except Blog.DoesNotExist:
        pass

    return redirect('viewArticles')

def viewDirectors(request):
    context = {}
    user = request.user
    userprofile = getattr(user, 'userprofile', None)

    club = None
    if userprofile:
        if userprofile.account_type == '2' and userprofile.director_profile:
            club = userprofile.director_profile.club

    if club:
        directors = UserProfile.objects.filter(
            account_type='2',
            director_profile__club=club
        ).select_related('user', 'director_profile')

        receptionists = UserProfile.objects.filter(
            account_type='5',
            receptionist_profile__club=club
        ).select_related('user', 'receptionist_profile')

        administrators = UserProfile.objects.filter(
            account_type='6',
            administrator_profile__club=club
        ).select_related('user', 'administrator_profile')

        accountants = UserProfile.objects.filter(
            account_type='7',
            accountant_profile__club=club
        ).select_related('user', 'accountant_profile')

        staff_list = []

        for director in directors:
            staff_list.append({
                'userprofile': director,
                'role': 'مدير عام',
                'role_en': 'General Manager',
                'profile': director.director_profile,
                'profile_type': 'director'
            })

        for receptionist in receptionists:
            staff_list.append({
                'userprofile': receptionist,
                'role': 'دعم فني',
                'role_en': 'technical support',
                'profile': receptionist.receptionist_profile,
                'profile_type': 'receptionist'
            })

        for administrator in administrators:
            staff_list.append({
                'userprofile': administrator,
                'role': 'إداري',
                'role_en': 'Administrator',
                'profile': administrator.administrator_profile,
                'profile_type': 'administrator'
            })

        for accountant in accountants:
            staff_list.append({
                'userprofile': accountant,
                'role': 'محاسب',
                'role_en': 'Accountant',
                'profile': accountant.accountant_profile,
                'profile_type': 'accountant'
            })

        staff_list.sort(key=lambda x: x['userprofile'].creation_date, reverse=True)

    else:
        staff_list = []

    context['LANGUAGE_CODE'] = translation.get_language()
    context['staff_list'] = staff_list
    return render(request, 'club_dashboard/directors/viewDirectors.html', context)

def addDirector(request):
    context = {}
    user = request.user

    if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
        messages.error(request, "Unauthorized access.")
        return redirect('club_dashboard_index')

    club = user.userprofile.director_profile.club

    ROLE_CHOICES = [
        ('2', 'مدير عام', 'General Manager'),
        ('5', 'دعم فني', 'Technical support'),
        ('6', 'إداري', 'Administrator'),
        ('7', 'محاسب', 'Accountant'),
    ]

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        role = request.POST.get('role')

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return redirect('addDirector')

        if User.objects.filter(email=email).exists():
            messages.error(request, "Email is already in use.")
            return redirect('addDirector')

        if role not in ['2', '5', '6', '7']:
            messages.error(request, "Invalid role selected.")
            return redirect('addDirector')

        form = None
        if role == '2':
            form = DirectorProfileForm(request.POST)
        elif role == '5':
            form = ReceptionistProfileForm(request.POST)
        elif role == '6':
            form = AdministratorProfileForm(request.POST)
        elif role == '7':
            form = AccountantProfileForm(request.POST)

        if form and form.is_valid():
            new_user = User.objects.create(username=username, email=email)
            if password:
                new_user.set_password(password)
            new_user.save()

            profile = form.save(commit=False)
            profile.club = club
            profile.save()

            user_profile_data = {
                'user': new_user,
                'account_type': role,
            }

            if role == '2':
                user_profile_data['director_profile'] = profile
            elif role == '5':
                user_profile_data['receptionist_profile'] = profile
            elif role == '6':
                user_profile_data['administrator_profile'] = profile
            elif role == '7':
                user_profile_data['accountant_profile'] = profile

            UserProfile.objects.create(**user_profile_data)

            role_names = {
                '2': 'Director',
                '5': 'Receptionist',
                '6': 'Administrator',
                '7': 'Accountant'
            }
            messages.success(request, f"{role_names[role]} added successfully.")
            return redirect('viewDirectors')
        else:
            messages.error(request, "Please correct the form errors.")

    context['LANGUAGE_CODE'] = translation.get_language()
    context['role_choices'] = ROLE_CHOICES
    return render(request, 'club_dashboard/directors/addDirector.html', context)


def editDirector(request, id, role):
    context = {}
    user = request.user

    if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
        messages.error(request, "Unauthorized access.")
        return redirect('club_dashboard_index')

    club = user.userprofile.director_profile.club

    ROLE_MAPPING = {
        '2': {
            'profile_model': DirectorProfile,
            'profile_field': 'director_profile',
            'form_class': DirectorProfileForm,
            'name': 'Director',
            'name_ar': 'مدير عام'
        },
        '5': {
            'profile_model': ReceptionistProfile,
            'profile_field': 'receptionist_profile',
            'form_class': ReceptionistProfileForm,
            'name': 'technical support',
            'name_ar': 'دعم فني'
        },
        '6': {
            'profile_model': AdministrativeProfile,
            'profile_field': 'administrator_profile',
            'form_class': AdministratorProfileForm,
            'name': 'Administrator',
            'name_ar': 'إداري'
        },
        '7': {
            'profile_model': AccountantProfile,
            'profile_field': 'accountant_profile',
            'form_class': AccountantProfileForm,
            'name': 'Accountant',
            'name_ar': 'محاسب'
        }
    }

    if role not in ROLE_MAPPING:
        messages.error(request, "Invalid role specified.")
        return redirect('viewDirectors')

    role_info = ROLE_MAPPING[role]

    try:
        profile = get_object_or_404(role_info['profile_model'], id=id, club=club)
        user_profile = get_object_or_404(UserProfile, **{role_info['profile_field']: profile})
        staff_user = user_profile.user
    except:
        messages.error(request, f"{role_info['name']} not found or unauthorized access.")
        return redirect('viewDirectors')

    form = role_info['form_class'](instance=profile)
    username = staff_user.username
    email = staff_user.email

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')

        form = role_info['form_class'](request.POST, instance=profile)

        if form.is_valid():
            if User.objects.exclude(id=staff_user.id).filter(username=username).exists():
                messages.error(request, "Username already exists.")
                return redirect('editDirector', id=id, role=role)

            if User.objects.exclude(id=staff_user.id).filter(email=email).exists():
                messages.error(request, "Email is already in use.")
                return redirect('editDirector', id=id, role=role)

            staff_user.username = username
            staff_user.email = email
            if password:
                staff_user.set_password(password)
            staff_user.save()

            updated_profile = form.save(commit=False)
            updated_profile.club = club
            updated_profile.save()

            messages.success(request, f"{role_info['name']} updated successfully.")
            return redirect('viewDirectors')
        else:
            messages.error(request, "Please correct the form errors.")

    context['LANGUAGE_CODE'] = translation.get_language()
    context['role'] = role
    context['role_name'] = role_info['name']
    context['role_name_ar'] = role_info['name_ar']

    return render(request, 'club_dashboard/directors/editDirector.html', {
        'form': form,
        'email': email,
        'username': username,
        'staff_user': staff_user,
        'profile': profile,
        'role': role,
        'role_name': role_info['name'],
        'role_name_ar': role_info['name_ar'],
        'LANGUAGE_CODE': context['LANGUAGE_CODE']
    })

def deleteDirector(request, id, role):
    user = request.user

    # ✅ Ensure the user is a director before proceeding
    if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
        messages.error(request, "Unauthorized access.")
        return redirect('viewDirectors')

    club = user.userprofile.director_profile.club

    # Role mapping for profile deletion
    ROLE_MAPPING = {
        '2': {
            'profile_model': DirectorProfile,
            'profile_field': 'director_profile',
            'name': 'Director',
            'name_ar': 'مدير عام'
        },
        '5': {
            'profile_model': ReceptionistProfile,
            'profile_field': 'receptionist_profile',
            'name': 'Receptionist',
            'name_ar': 'موظف استقبال'
        },
        '6': {
            'profile_model': AdministrativeProfile,
            'profile_field': 'administrator_profile',
            'name': 'Administrator',
            'name_ar': 'إداري'
        },
        '7': {
            'profile_model': AccountantProfile,
            'profile_field': 'accountant_profile',
            'name': 'Accountant',
            'name_ar': 'محاسب'
        }
    }

    # Validate role
    if role not in ROLE_MAPPING:
        messages.error(request, "Invalid role specified.")
        return redirect('viewDirectors')

    role_info = ROLE_MAPPING[role]

    # Get the profile and user based on role
    try:
        staff_profile = get_object_or_404(role_info['profile_model'], id=id, club=club)
        user_profile = get_object_or_404(UserProfile, **{role_info['profile_field']: staff_profile})
        staff_user = user_profile.user
    except:
        messages.error(request, f"{role_info['name']} not found or unauthorized access.")
        return redirect('viewDirectors')

    # ✅ Additional security check - ensure staff belongs to the same club
    if staff_profile.club != club:
        messages.error(request, f"You cannot delete a {role_info['name'].lower()} from another club.")
        return redirect('viewDirectors')

    # ✅ Prevent directors from deleting themselves
    if staff_user == user:
        messages.error(request, "You cannot delete your own account.")
        return redirect('viewDirectors')

    # ✅ Delete the profile and user
    try:
        staff_profile.delete()
        staff_user.delete()
        messages.success(request, f"{role_info['name']} deleted successfully.")
    except Exception as e:
        messages.error(request, f"An error occurred while deleting the {role_info['name'].lower()}.")

    return redirect('viewDirectors')

def viewClubNotifications(request):
    context = {}
    """Displays all club notifications and marks them as read."""
    user = request.user

    # ✅ Ensure the user has a valid director profile
    # if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
    #     messages.error(request, "Unauthorized access.")
    #     return redirect('home')

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)


    # ✅ Fetch all notifications for the club
    notifications = Notification.objects.filter(club=club).order_by('-created_at')

    # ✅ Ensure only unread notifications are marked as read
    unread_count = notifications.filter(is_read=False).update(is_read=True)
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/notifications/viewClubNotifications.html', {
        'notifications': notifications,
        'unread_count': unread_count,  # ✅ Pass unread count for better UI
        'club':club,
    })


def delete_notification(request, notification_id):
    """Delete a specific notification"""
    if request.method == 'POST':
        try:
            notification = Notification.objects.get(id=notification_id)
            # Check if the notification belongs to the user's club
            club = getattr(request.user.userprofile.director_profile, 'club', None) or getattr(request.user.userprofile.administrator_profile, 'club', None)
            if notification.club == club:
                notification.delete()
                messages.success(request, "Notification deleted successfully.")
            else:
                messages.error(request, "You don't have permission to delete this notification.")
        except Notification.DoesNotExist:
            messages.error(request, "Notification not found.")

    return redirect('viewClubNotifications')

def delete_all_notifications(request):
    """Delete all notifications for the club"""
    if request.method == 'POST':
        club = getattr(request.user.userprofile.director_profile, 'club', None) or getattr(request.user.userprofile.administrator_profile, 'club', None)
        if club:
            deleted_count, _ = Notification.objects.filter(club=club).delete()
            messages.success(request, f"Deleted {deleted_count} notifications.")
        else:
            messages.error(request, "No club associated with your account.")

    return redirect('viewClubNotifications')


def mark_notifications_read(request):
    """Marks all unread notifications as read for the club."""
    user = request.user

    # ✅ Ensure the user has a valid director profile
    # if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
    #     return JsonResponse({'status': 'error', 'message': 'Unauthorized access'}, status=403)

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)


    # ✅ Mark only unread notifications as read
    updated_count = Notification.objects.filter(club=club, is_read=False).update(is_read=True)

    return JsonResponse({'status': 'success', 'message': f'Marked {updated_count} notifications as read'})


@login_required
def salon_appointments(request):
    context = {}
    user = request.user
    days = ['السبت', 'الأحد', 'الإثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']

    from datetime import time

    time_slots = []
    for hour in range(12, 24):
        time_slots.append(time(hour, 0))
    time_slots.append(time(0, 0))

    schedule = {}

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)
    if not club:
        messages.error(request, "No club assigned to your profile. Please contact an administrator.")
        return render(request, 'student/blog/viewArticles.html', {
            'schedule': {},
            'days': days,
            'time_slots': [slot.strftime('%I:%M %p') for slot in time_slots]
        })

    for day in days:
        schedule[day] = []
        appointments = SalonAppointment.objects.filter(
            day=day,
            club=club,
            is_paid=True
        ).order_by('start_time')

        for slot in time_slots:
            slot_end = (datetime.combine(datetime.today(), slot) + timedelta(hours=1)).time()

            slot_appointments = appointments.filter(
                start_time__gte=slot,
                start_time__lt=slot_end
            )

            booking_count = slot_appointments.count()

            slot_info = {
                'time': slot.strftime('%I:%M %p'),
                'booking_count': booking_count,
                'has_bookings': booking_count > 0,
                'appointments': [
                    {
                        'id': appt.id,
                        'start': appt.start_time.strftime('%I:%M %p'),
                        'end': appt.end_time.strftime('%I:%M %p') if appt.end_time else "N/A"
                    } for appt in slot_appointments if hasattr(appt, 'booking')
                ]
            }

            schedule[day].append(slot_info)
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/salon_appointments.html', {
        'schedule': schedule,
        'days': days,
        'time_slots': [slot.strftime('%I:%M %p') for slot in time_slots],
        'club': club,
    })


@login_required
def slot_appointments(request, day, time):
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)
    if not club:
        messages.error(request, "No club assigned to your profile. Please contact an administrator.")
        return redirect('index')

    try:
        if ':' in time:
            hour_str, minute_str = time.split(':')
            hour = int(hour_str)
            minute = int(minute_str.split(' ')[0])
            period = time.split(' ')[1]

            if period == 'PM' and hour < 12:
                hour += 12
            elif period == 'AM' and hour == 12:
                hour = 0

            time_obj = datetime.strptime(f'{hour:02d}:{minute:02d}:00', '%H:%M:%S').time()
            slot_end = (datetime.combine(datetime.today(), time_obj) + timedelta(hours=1)).time()

            # Only get appointments with confirmed payments
            appointments = SalonAppointment.objects.filter(
                day=day,
                club=club,
                is_paid=True,
                start_time__gte=time_obj,
                start_time__lt=slot_end
            ).select_related('booking')

            appointment_details = []
            for appt in appointments:
                if hasattr(appt, 'booking'):
                    booking = appt.booking
                    services = BookingService.objects.filter(booking=booking).select_related('service')
                    service_names = ", ".join([s.service.title for s in services])
                    total_price = sum(s.service.price for s in services)

                    appointment_details.append({
                        'id': appt.id,
                        'services': service_names,
                        'employee': booking.employee,
                        'start_time': appt.start_time.strftime('%I:%M %p'),
                        'end_time': appt.end_time.strftime('%I:%M %p'),
                        'total_price': total_price
                    })

            context = {
                'day': day,
                'time_slot': time,
                'appointments': appointment_details,
                'club':club,
            }
            context['LANGUAGE_CODE'] = translation.get_language()
            return render(request, 'club_dashboard/slot_appointments.html', context)

    except Exception as e:
        messages.error(request, f"Error retrieving appointments: {str(e)}")
        return redirect('club_salon_appointments')

@login_required
def appointment_details(request, appointment_id):
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    if not club:
        messages.error(request, "لم يتم تحديد نادٍ لهذا المستخدم.")
        return redirect('index')

    appointment = get_object_or_404(SalonAppointment, id=appointment_id)
    if appointment.club != club:
        messages.error(request, "ليس لديك صلاحية لعرض هذا الموعد.")
        return redirect('receptionist_salon_appointments')

    try:
        booking = appointment.booking
        # Explicitly fetch booking services
        booking_services = BookingService.objects.filter(booking=booking).select_related('service')

        if not booking_services.exists():
            messages.warning(request, "لا توجد خدمات مرتبطة بهذا الحجز")

        # For debugging
        print(f"Found {booking_services.count()} services for booking {booking.id}")

        context = {
            'appointment': appointment,
            'booking': booking,
            'employee': booking.employee,
            'booking_services': booking_services,  # Change the variable name to be more explicit
            'total_price': sum(bs.service.price for bs in booking_services),
            'total_duration': sum(bs.service.duration for bs in booking_services),
            'club':club,
            'created_by_type': getattr(booking, 'created_by_type', 'unknown'),
            'created_by_name': getattr(booking, 'created_by_name', 'غير محدد'),
        }
    except Exception as e:
        messages.error(request, f"لم يتم العثور على معلومات الحجز: {str(e)}")
        return redirect('club_salon_appointments')
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/club_appointment_details.html', context)

def reviews_list(request):
    context = {}
    """
    Fetch all reviews for the club associated with the logged-in user.
    """
    # Assuming the logged-in user is a director or admin
    user = request.user

    try:
        # Get the club associated with the logged-in user
        club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)


        # Fetch all reviews for coaches in this club
        reviews = Review.objects.filter(coach__club=club).select_related(
            'student', 'coach'
        ).order_by('-created_at')
        # Calculate average rating for the club
        avg_rating = reviews.aggregate(Avg('rating'))['rating__avg'] or 0
        context['LANGUAGE_CODE'] = translation.get_language()
        return render(request, 'club_dashboard/reviews_list.html', {
            'reviews': reviews,
            'avg_rating': avg_rating,
            'total_reviews': reviews.count(),
            'club':club
        })

    except AttributeError:
        messages.error(request, "لا يمكن العثور على النادي الخاص بك.")
        return redirect('club_dashboard')


def viewReceptionists(request):
    context = {}
    """Displays a list of coaches belonging to the club."""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)
    receptionist_userprofile = UserProfile.objects.filter(
        account_type='5',
        receptionist_profile__club=getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    )
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/receptionists/viewReceptionists.html', {'receptionist_userprofile': receptionist_userprofile,'club':club})

@login_required
def addReceptionist(request):
    context = {}
    user = request.user

    # if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
    #     messages.error(request, "Unauthorized access.")
    #     return redirect('home')

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)


    # Initialize the correct form
    form = ReceptionistSignupForm(initial={'club': club})

    if request.method == 'POST':
        form = ReceptionistSignupForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']

            if User.objects.filter(username=username).exists():
                messages.error(request, "Username already exists.")
                return redirect('addReceptionist')

            if User.objects.filter(email=email).exists():
                messages.error(request, "Email is already in use.")
                return redirect('addReceptionist')

            receptionist = User.objects.create(username=username, email=email)
            if password:
                receptionist.set_password(password)
            receptionist.save()

            receptionist_profile = ReceptionistProfile.objects.create(
                full_name=form.cleaned_data['full_name'],
                phone=form.cleaned_data['phone'],
                email=email,
                club=form.cleaned_data['club'],  # Get club from form
                about=form.cleaned_data.get('about', "")
            )

            UserProfile.objects.create(user=receptionist, account_type='5', receptionist_profile=receptionist_profile)

            send_notification(user, club, f" موظف الدعم الفني الجديد 📢 {username} انضم إلى {club.name}.")

            messages.success(request, "Receptionist added successfully.")
            return redirect('viewReceptionists')
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/receptionists/addReceptionist.html', {'form': form,'club':club})


@login_required
def editReceptionist(request, id):
    context = {}
    """Edits an existing receptionist's details."""
    user = request.user

    # if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
    #     messages.error(request, "Unauthorized access.")
    #     return redirect('home')

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    receptionist_profile = get_object_or_404(ReceptionistProfile, id=id)
    receptionist = get_object_or_404(User, userprofile__receptionist_profile=receptionist_profile)

    form = ReceptionistProfileForm(instance=receptionist_profile)

    if request.method == 'POST':
        new_username = request.POST.get('username')
        new_email = request.POST.get('email')
        password = request.POST.get('password')

        form = ReceptionistProfileForm(request.POST, instance=receptionist_profile)
        if form.is_valid():
            username_changed = new_username != receptionist.username
            email_changed = new_email != receptionist.email

            receptionist.username = new_username
            receptionist.email = new_email
            if password:
                receptionist.set_password(password)
            receptionist.save()

            form.save()

            notification_message = f" تم تحديث ملف موظف الدعم الفني 📝 {receptionist.username}."
            if username_changed or email_changed:
                notification_message += " (تم تعديل اسم المستخدم/البريد الإلكتروني.)"

            send_notification(user, club, notification_message)

            messages.success(request, "Receptionist profile updated successfully.")
            return redirect('viewReceptionists')
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/receptionists/editReceptionist.html', {
        'form': form,
        'receptionist': receptionist,
        'club':club
    })


@login_required
def deleteReceptionist(request, id):
    """Deletes a receptionist from the club."""
    user = request.user

    # if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
    #     messages.error(request, "Unauthorized access.")
    #     return redirect('home')

    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    receptionist_profile = get_object_or_404(ReceptionistProfile, id=id)
    receptionist = get_object_or_404(User, userprofile__receptionist_profile=receptionist_profile)

    receptionist_name = receptionist.username

    receptionist_profile.delete()
    receptionist.delete()

    send_notification(user, club, f" تم حذف موظف الدعم الفني 🗑️ {receptionist_name} .من المنصة ")

    messages.success(request, "Receptionist has been deleted successfully.")
    return redirect('viewReceptionists')

@login_required
def cancel_appointment(request, appointment_id):
    club = get_user_club(request.user)

    if not club:
        messages.error(request, "لم يتم تحديد نادٍ لهذا المستخدم.")
        return redirect('index')

    appointment = get_object_or_404(SalonAppointment, id=appointment_id)
    if appointment.club != club:
        messages.error(request, "ليس لديك صلاحية لإلغاء هذا الموعد.")
        return redirect('club_salon_appointments')

    try:
        booking = appointment.booking
        BookingService.objects.filter(booking=booking).delete()
        booking.delete()
        appointment.delete()

        messages.success(request, "تم إلغاء الحجز بنجاح")
    except:
        messages.error(request, "لم يتم العثور على حجز لهذا الموعد")

    return redirect('club_salon_appointments')

@login_required
def club_orders(request):
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    orders = Order.objects.filter(club=club).order_by('-created_at')

    status_filter = request.GET.get('status')
    if status_filter and status_filter != 'all':
        orders = orders.filter(status=status_filter)

    from django.db.models import Count, Q, Case, When, IntegerField, Value, CharField, Sum, F, DecimalField

    orders = orders.annotate(
        has_products=Count('items', filter=Q(items__product__isnull=False)),
        has_services=Count('items', filter=Q(items__service__isnull=False))
    ).annotate(
        order_type=Case(
            When(has_products__gt=0, has_services=0, then=Value('products')),
            When(has_products=0, has_services__gt=0, then=Value('services')),
            When(has_products__gt=0, has_services__gt=0, then=Value('mixed')),
            default=Value('unknown'),
            output_field=CharField()
        ),
        order_type_display=Case(
            When(has_products__gt=0, has_services=0, then=Value('منتجات')),
            When(has_products=0, has_services__gt=0, then=Value('خدمات')),
            When(has_products__gt=0, has_services__gt=0, then=Value('منتجات وخدمات')),
            default=Value('غير معروف'),
            output_field=CharField()
        )
    )

    # Add commission summary statistics
    from django.db.models import Sum
    total_orders_value = orders.aggregate(total=Sum('total_price'))['total'] or 0
    total_commissions = orders.aggregate(total=Sum('total_vendor_commission'))['total'] or 0
    total_club_revenue = orders.aggregate(total=Sum('club_revenue'))['total'] or 0

    context = {
        'orders': orders,
        'status_filter': status_filter or 'all',
        'club': club,
        'commission_stats': {
            'total_orders_value': total_orders_value,
            'total_commissions': total_commissions,
            'total_club_revenue': total_club_revenue
        }
    }
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/orders/club_orders.html', context)

@login_required
def update_order_status(request, order_id):
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    if not club:
        return JsonResponse({'status': 'error', 'message': 'Unauthorized access'}, status=403)

    if request.method == 'POST':
        try:
            print(f"Request to update order {order_id} - POST data: {request.POST}")

            order = get_object_or_404(Order, id=order_id, club=club)
            new_status = request.POST.get('status')

            # Handle cancellation with reasons
            if new_status == 'cancelled':
                return handle_order_cancellation(request, order, user)

            # Handle other status updates
            return handle_regular_status_update(request, order, new_status, user)

        except Order.DoesNotExist:
            print(f"Order {order_id} not found")
            return JsonResponse({'status': 'error', 'message': 'Order not found'}, status=404)
        except Exception as e:
            print(f"Unexpected error in update_order_status: {str(e)}")
            return JsonResponse({'status': 'error', 'message': f'Server error: {str(e)}'}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)

def handle_order_cancellation(request, order, user):
    """Handle order cancellation with detailed reasons"""
    try:
        # Parse JSON data if it's sent as JSON
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST

        reason = data.get('cancellation_reason') or data.get('reason')
        custom_reason = data.get('custom_reason', '')
        additional_notes = data.get('additional_notes', '')

        if not reason:
            return JsonResponse({
                'status': 'error',
                'message': 'Cancellation reason is required'
            }, status=400)

        # Validate reason
        valid_reasons = [choice[0] for choice in OrderCancellation.CANCELLATION_REASONS]
        if reason not in valid_reasons:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid cancellation reason'
            }, status=400)

        # If reason is 'other', custom_reason is required
        if reason == 'other' and not custom_reason.strip():
            return JsonResponse({
                'status': 'error',
                'message': 'Custom reason is required when selecting "Other"'
            }, status=400)

        # Check if order can be cancelled
        if order.status in ['completed', 'cancelled']:
            return JsonResponse({
                'status': 'error',
                'message': f'Cannot cancel order with status: {order.get_status_display()}'
            }, status=400)

        old_status = order.status

        # Create cancellation record first
        cancellation = OrderCancellation.objects.create(
            order=order,
            reason=reason,
            custom_reason=custom_reason if reason == 'other' else '',
            additional_notes=additional_notes,
            cancelled_by=user
        )

        # Update order status
        order.status = 'cancelled'
        order.save()

        # Handle stock restoration for confirmed orders
        if old_status == 'confirmed':
            restore_product_stock(order)

        # Handle service cancellations
        handle_service_cancellation(order)

        # Create notification
        create_cancellation_notification(order, cancellation)

        print(f"Order {order.id} cancelled successfully by user {user.id}")

        return JsonResponse({
            'status': 'success',
            'message': f'تم إلغاء الطلب #{order.id} بنجاح',
            'cancellation_reason': cancellation.get_reason_display_text()
        })

    except Exception as e:
        print(f"Error in handle_order_cancellation: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'status': 'error',
            'message': f'Error cancelling order: {str(e)}'
        }, status=500)

def handle_regular_status_update(request, order, new_status, user):
    """Handle regular status updates (non-cancellation)"""
    try:
        print(f"Current order status: {order.status}, New status: {new_status}")

        if new_status not in dict(Order.STATUS_CHOICES):
            return JsonResponse({'status': 'error', 'message': 'Invalid status'}, status=400)

        old_status = order.status
        order.status = new_status

        try:
            order.full_clean()
            order.save()
            print(f"Order saved successfully with new status: {new_status}")
        except Exception as save_error:
            print(f"Error saving order: {save_error}")
            return JsonResponse({'status': 'error', 'message': f'Database error: {str(save_error)}'}, status=500)

        # Determine order type
        try:
            has_products = OrderItem.objects.filter(order=order, product__isnull=False).exists()
            has_services = OrderItem.objects.filter(order=order, service__isnull=False).exists()
            order_type = "mixed" if (has_products and has_services) else "products" if has_products else "services"
        except Exception:
            order_type = "unknown"

        # Handle confirmation logic
        if new_status == 'confirmed' and old_status == 'pending':
            process_order_confirmation(order)

        # Create notification
        create_status_update_notification(order, new_status, order_type)

        return JsonResponse({
            'status': 'success',
            'message': f'تم تحديث حالة الطلب إلى {dict(Order.STATUS_CHOICES)[new_status]}'
        })

    except Exception as e:
        print(f"Error in handle_regular_status_update: {str(e)}")
        return JsonResponse({'status': 'error', 'message': f'Server error: {str(e)}'}, status=500)

def restore_product_stock(order):
    """Restore product stock when cancelling confirmed orders"""
    try:
        product_items = OrderItem.objects.filter(order=order, product__isnull=False)
        for item in product_items:
            product = item.product
            product.stock += item.quantity
            product.save()
            print(f"Restored product stock for {product.title}: {product.stock}")
    except Exception as e:
        print(f"Error restoring product stock: {str(e)}")

def handle_service_cancellation(order):
    """Handle service-related cancellations"""
    try:
        service_items = OrderItem.objects.filter(order=order, service__isnull=False)
        for item in service_items:
            service = item.service

            # Cancel active service orders
            try:
                service_orders = ServiceOrderModel.objects.filter(
                    service=service,
                    student=order.user,
                    is_complited=False
                )

                for service_order in service_orders:
                    service_order.is_complited = True  # Mark as completed (cancelled)
                    service_order.save()
                    print(f"Cancelled service order: {service_order.id}")

            except Exception as e:
                print(f"Error cancelling service orders: {str(e)}")

            # Handle appointment cancellations
            try:
                if hasattr(order.user, 'userprofile') and hasattr(order.user.userprofile, 'student_profile'):
                    student_profile = order.user.userprofile.student_profile

                    bookings = SalonBooking.objects.filter(
                        student=student_profile
                    ).select_related('appointment')

                    booking_services = BookingService.objects.filter(
                        service=service,
                        booking__in=bookings
                    )

                    for booking_service in booking_services:
                        if (booking_service.booking and
                                hasattr(booking_service.booking, 'appointment') and
                                booking_service.booking.appointment.is_paid):

                            appointment = booking_service.booking.appointment
                            appointment.is_paid = False  # Revert payment status
                            appointment.save()
                            print(f"Reverted payment for appointment ID {appointment.id}")

            except Exception as e:
                print(f"Error handling appointment cancellations: {str(e)}")

    except Exception as e:
        print(f"Error in handle_service_cancellation: {str(e)}")



import logging
from django.db import transaction
from django.utils import timezone
from django.contrib.auth.models import User

logger = logging.getLogger(__name__)
def process_order_confirmation(order):
    """
    Process order confirmation - update stock and subscriptions
    This should be called when an order status changes from 'pending' to 'confirmed'
    """
    try:
        with transaction.atomic():
            # Process product items - reduce stock
            product_items = order.items.filter(product__isnull=False)
            for item in product_items:
                product = item.product
                if product.stock >= item.quantity:
                    product.stock -= item.quantity
                    product.save()
                    logger.info(f"Updated product stock for {product.title}: {product.stock}")
                else:
                    raise ValueError(f"Insufficient stock for product {product.title}")

            # Process service items - update subscriptions
            service_items = order.items.filter(service__isnull=False)
            for item in service_items:
                service = item.service

                # Handle service subscriptions
                from students.models import ServiceOrderModel

                # Look for any active subscription for this user and service
                existing_service_order = ServiceOrderModel.objects.filter(
                    student=order.user,
                    service=service,
                    is_complited=False  # Only active subscriptions
                ).order_by('-end_datetime').first()

                if existing_service_order:
                    # Extend existing subscription
                    subscription_months = service.pricing_period_months * item.quantity

                    # If the current subscription is still active, extend from its end date
                    # Otherwise, extend from now
                    if existing_service_order.end_datetime > timezone.now():
                        new_end_datetime = existing_service_order.end_datetime + timezone.timedelta(days=subscription_months * 30)
                    else:
                        new_end_datetime = timezone.now() + timezone.timedelta(days=subscription_months * 30)

                    # Update the existing subscription
                    existing_service_order.end_datetime = new_end_datetime
                    existing_service_order.price += service.price * item.quantity
                    existing_service_order.creation_date = timezone.now()
                    existing_service_order.is_complited = False  # Ensure it's still active
                    existing_service_order.save()

                    logger.info(f"Extended existing subscription for service {service.title} (ID: {service.id}) until {new_end_datetime}")

                else:
                    # Create new service subscription
                    subscription_months = service.pricing_period_months * item.quantity
                    end_datetime = timezone.now() + timezone.timedelta(days=subscription_months * 30)

                    new_service_order = ServiceOrderModel.objects.create(
                        service=service,
                        student=order.user,
                        price=service.price * item.quantity,
                        is_complited=False,
                        end_datetime=end_datetime,
                        creation_date=timezone.now()
                    )

                    logger.info(f"Created new subscription for service {service.title} (ID: {service.id}) until {end_datetime}")

                # Handle appointment payments
                try:
                    from receptionist_dashboard.models import SalonBooking, BookingService

                    # Find student profile
                    student_profile = None
                    if hasattr(order.user, 'userprofile') and hasattr(order.user.userprofile, 'student_profile'):
                        student_profile = order.user.userprofile.student_profile

                    if student_profile:
                        bookings = SalonBooking.objects.select_related('appointment')

                        booking_services = BookingService.objects.filter(
                            service=service,
                            booking__in=bookings
                        )

                        for booking_service in booking_services:
                            try:
                                if booking_service.booking and hasattr(booking_service.booking, 'appointment'):
                                    appointment = booking_service.booking.appointment
                                    if not appointment.is_paid:
                                        appointment.is_paid = True
                                        appointment.save()
                                        logger.info(f"Updated appointment ID {appointment.id} to paid")
                            except Exception as appt_error:
                                logger.error(f"Error updating appointment: {str(appt_error)}")

                except ImportError:
                    logger.warning("Salon booking models not available")
                except Exception as e:
                    logger.error(f"Error handling appointment payments: {str(e)}")

            # NEW: Process commission tracking
            process_order_commissions(order)

            # Update order commission fields
            order.update_commission_fields()


            logger.info(f"Successfully processed order confirmation for order {order.id}")
            return True

    except Exception as e:
        logger.error(f"Error processing order confirmation for order {order.id}: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False




def process_order_commissions(order):
    """Process and save vendor commissions for confirmed order"""
    try:
        # Clear existing commission records
        OrderVendorCommission.objects.filter(order=order).delete()

        # Get commission breakdown
        breakdown = order.calculate_commission_breakdown()

        # Create commission records for each vendor
        for vendor_data in breakdown['vendor_breakdowns']:
            OrderVendorCommission.objects.create(
                order=order,
                vendor=vendor_data['vendor'],
                total_amount=vendor_data['total_amount'],
                commission_rate=vendor_data['commission_rate'],
                commission_amount=vendor_data['commission_amount']
            )

        logger.info(f"Created commission records for order {order.id}")

    except Exception as e:
        logger.error(f"Error processing commissions for order {order.id}: {str(e)}")





def create_cancellation_notification(order, cancellation):
    """Create notification for order cancellation"""
    try:
        # Determine order type for notification message
        has_products = OrderItem.objects.filter(order=order, product__isnull=False).exists()
        has_services = OrderItem.objects.filter(order=order, service__isnull=False).exists()

        if has_products and not has_services:
            order_type_ar = "المنتجات"
        elif has_services and not has_products:
            order_type_ar = "الخدمات"
        else:
            order_type_ar = "المنتجات والخدمات"

        reason_text = cancellation.get_reason_display_text()

        message = (f"تم إلغاء طلب {order_type_ar} رقم #{order.id}. "
                   f"سبب الإلغاء: {reason_text}. "
                   f"يرجى التواصل مع خدمة العملاء للمزيد من المعلومات.")

        Notification.objects.create(
            user=order.user,
            message=message,
            notification_type='order_cancelled'
        )
        print(f"Created cancellation notification for user {order.user.id}")

    except Exception as e:
        print(f"Error creating cancellation notification: {e}")

def create_status_update_notification(order, new_status, order_type):
    """Create notification for regular status updates"""
    try:
        if new_status == 'confirmed':
            if order_type == "products":
                message = f"تم تأكيد طلب المنتجات رقم #{order.id} وسيتم تجهيزه قريباً."
            elif order_type == "services":
                message = f"تم تأكيد طلب الخدمات رقم #{order.id} وسيتم تفعيلها قريباً."
            else:
                message = f"تم تأكيد طلب المنتجات والخدمات رقم #{order.id} وسيتم معالجته قريباً."
        elif new_status == 'completed':
            if order_type == "products":
                message = f"تم توصيل منتجاتك من الطلب رقم #{order.id} بنجاح. شكراً لك!"
            elif order_type == "services":
                message = f"تم تفعيل خدماتك من الطلب رقم #{order.id} بنجاح. شكراً لك!"
            else:
                message = f"تم اكتمال طلب المنتجات والخدمات رقم #{order.id} بنجاح. شكراً لك!"
        else:
            message = f"تم تحديث حالة الطلب رقم #{order.id} إلى {dict(Order.STATUS_CHOICES)[new_status]}."

        Notification.objects.create(
            user=order.user,
            message=message,
            notification_type='order_update'
        )
        print(f"Created notification for user {order.user.id}")

    except Exception as e:
        print(f"Error creating notification: {e}")

@login_required
def order_details_api(request, order_id):
    user = request.user
    print(f"طلب تفاصيل الطلب {order_id} بواسطة {user}")

    try:
        club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

        order = Order.objects.get(id=order_id, club=club)
        order_items = OrderItem.objects.filter(order=order)

        # Get cancellation details if order is cancelled
        cancellation = None
        if order.status == 'cancelled':
            try:
                cancellation = OrderCancellation.objects.get(order=order)
            except OrderCancellation.DoesNotExist:
                print(f"No cancellation record found for cancelled order {order_id}")

        # Get commission breakdown
        commission_breakdown = None
        vendor_commissions = None
        if order.status == 'confirmed':
            commission_breakdown = order.calculate_commission_breakdown()
            # Convert Decimal to float for JSON serialization
            commission_breakdown['total_vendor_commission'] = float(commission_breakdown['total_vendor_commission'])
            commission_breakdown['club_revenue'] = float(commission_breakdown['club_revenue'])
            for vendor in commission_breakdown['vendor_breakdowns']:
                vendor['total_amount'] = float(vendor['total_amount'])
                vendor['commission_amount'] = float(vendor['commission_amount'])

            # Get vendor commission records
            vendor_commissions = OrderVendorCommission.objects.filter(order=order).select_related('vendor')

        context = {
            'order': order,
            'order_items': order_items,
            'cancellation': cancellation,
            'has_transfer_receipt': bool(order.transfer_receipt),
            'transfer_receipt_url': order.transfer_receipt.url if order.transfer_receipt else None,
            'commission_breakdown': commission_breakdown,
            'vendor_commissions': vendor_commissions,
        }
        context['LANGUAGE_CODE'] = translation.get_language()
        html = render_to_string('club_dashboard/orders/order_details_modal.html', context)

        return JsonResponse({
            'status': 'success',
            'html': html
        })
    except Order.DoesNotExist:
        print(f"الطلب {order_id} غير موجود")
        return JsonResponse({'status': 'error', 'message': 'Order not found'}, status=404)
    except Exception as e:
        print(f"خطأ غير متوقع: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
def vendor_commission_report(request):
    """Generate vendor commission report"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    if not club:
        return JsonResponse({'status': 'error', 'message': 'Unauthorized access'}, status=403)

    # Get date filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Base queryset
    commissions = OrderVendorCommission.objects.filter(
        order__club=club,
        order__status='confirmed'
    ).select_related('vendor', 'order')

    # Apply date filters
    if start_date:
        commissions = commissions.filter(order__created_at__gte=start_date)
    if end_date:
        commissions = commissions.filter(order__created_at__lte=end_date)

    # Group by vendor
    from django.db.models import Sum, Count
    vendor_summary = commissions.values(
        'vendor__id',
        'vendor__business_name',
        'vendor__full_name'
    ).annotate(
        total_orders=Count('order', distinct=True),
        total_sales=Sum('total_amount'),
        total_commission=Sum('commission_amount'),
        avg_commission_rate=models.Avg('commission_rate')
    ).order_by('-total_commission')

    context = {
        'vendor_summary': vendor_summary,
        'commissions': commissions,
        'start_date': start_date,
        'end_date': end_date,
        'club': club
    }

    return render(request, 'club_dashboard/reports/vendor_commission_report.html', context)



from students.models import OrderVendorCommission
@login_required
def order_full_details(request, order_id):
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    try:
        # Optimize queries by selecting related objects
        order = Order.objects.get(id=order_id, club=club)
        order_items = OrderItem.objects.filter(order=order).select_related(
            'product',
            'service',
            'product__creator__userprofile__Coach_profile',
            'service__creator__userprofile__Coach_profile'
        )

        # Get cancellation details if order is cancelled
        cancellation = None
        if order.status == 'cancelled':
            cancellation = OrderCancellation.objects.filter(order=order).first()

        # Get commission breakdown
        commission_breakdown = order.calculate_commission_breakdown()

        # Calculate VAT (15%)
        orderbeforevat = float(order.total_price)/ 1.15
        commission_breakdown['vat_amount'] = float(orderbeforevat) * 0.15

        # Organize items by vendor
        vendor_items = {}
        for item in order_items:
            vendor = None
            if item.product and hasattr(item.product.creator.userprofile, 'Coach_profile'):
                vendor = item.product.creator.userprofile.Coach_profile
            elif item.service and hasattr(item.service.creator.userprofile, 'Coach_profile'):
                vendor = item.service.creator.userprofile.Coach_profile

            if vendor:
                if vendor.id not in vendor_items:
                    vendor_items[vendor.id] = {
                        'vendor': vendor,
                        'items': []
                    }
                vendor_items[vendor.id]['items'].append(item)

        # Add items to vendor breakdowns if they exist
        if 'vendor_breakdowns' in commission_breakdown:
            for vendor_data in commission_breakdown['vendor_breakdowns']:
                vendor_id = vendor_data['vendor'].id
                if vendor_id in vendor_items:
                    vendor_data['items'] = vendor_items[vendor_id]['items']

        context = {
            'order': order,
            'order_items': order_items,
            'cancellation': cancellation,
            'has_transfer_receipt': bool(order.transfer_receipt),
            'transfer_receipt_url': order.transfer_receipt.url if order.transfer_receipt else None,
            'commission_breakdown': commission_breakdown,
            'club': club,
        }
        context['LANGUAGE_CODE'] = translation.get_language()

        return render(request, 'club_dashboard/orders/order_full_details.html', context)

    except Order.DoesNotExist:
        messages.error(request, _('Order not found'))
        return redirect('club_orders')
    except Exception as e:
        messages.error(request, _('An error occurred while loading order details'))
        logger.error(f"Error in order_full_details: {str(e)}")
        return redirect('club_orders')




@login_required
def get_cancellation_details(request, order_id):
    """API endpoint to get cancellation details for an order"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    if not club:
        return JsonResponse({'status': 'error', 'message': 'Unauthorized access'}, status=403)

    try:
        order = get_object_or_404(Order, id=order_id, club=club)

        if not hasattr(order, 'cancellation'):
            return JsonResponse({'status': 'error', 'message': 'Order is not cancelled'}, status=404)

        cancellation = order.cancellation

        return JsonResponse({
            'status': 'success',
            'cancellation': {
                'reason': cancellation.reason,
                'reason_display': cancellation.get_reason_display_text(),
                'custom_reason': cancellation.custom_reason,
                'additional_notes': cancellation.additional_notes,
                'cancelled_by': cancellation.cancelled_by.get_full_name() or cancellation.cancelled_by.username,
                'cancelled_at': cancellation.cancelled_at.strftime('%Y-%m-%d %H:%M:%S')
            }
        })

    except Exception as e:
        print(f"Error getting cancellation details: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def edit_shipment(request, shipment_id):
    """Edit a product shipment"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    try:
        shipment = ProductShipment.objects.get(id=shipment_id, product__club=club)
    except ProductShipment.DoesNotExist:
        messages.error(request, 'الشحنة غير موجودة!' if translation.get_language() == 'ar' else 'Shipment not found!')
        return redirect('viewProducts')

    if request.method == 'POST':
        form = ProductShipmentForm(request.POST, instance=shipment, club=club)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تحديث الشحنة بنجاح!' if translation.get_language() == 'ar' else 'Shipment updated successfully!')
            return redirect('view_product_shipments', product_id=shipment.product.id)
        else:
            messages.error(request, 'يرجى تصحيح الأخطاء أدناه.' if translation.get_language() == 'ar' else 'Please correct the errors below.')
    else:
        form = ProductShipmentForm(instance=shipment, club=club)

    context = {
        'form': form,
        'shipment': shipment,
        'product': shipment.product,
        'club': club,
        'LANGUAGE_CODE': translation.get_language(),
        'is_edit': True,
    }

    return render(request, 'club_dashboard/products/ProductsStock/add_edit_shipment.html', context)


def delete_shipment(request, shipment_id):
    """Delete a product shipment"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    try:
        shipment = ProductShipment.objects.get(id=shipment_id, product__club=club)
        product_id = shipment.product.id
        product_title = shipment.product.title
        quantity = shipment.quantity

        shipment.delete()

        messages.success(
            request,
            f'تم حذف شحنة المنتج "{product_title}" بكمية {quantity} وحدة بنجاح!'
            if translation.get_language() == 'ar'
            else f'Shipment for product "{product_title}" with quantity {quantity} units deleted successfully!'
        )

        return redirect('view_product_shipments', product_id=product_id)

    except ProductShipment.DoesNotExist:
        messages.error(request, 'الشحنة غير موجودة!' if translation.get_language() == 'ar' else 'Shipment not found!')
        return redirect('viewProducts')


# Update your existing add_shipment view to use the same template
def add_shipment(request):
    """Add a new product shipment"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    # Get product_id from URL parameters if provided
    product_id = request.GET.get('product_id')
    product = None

    if product_id:
        try:
            product = ProductsModel.objects.get(id=product_id, club=club)
        except ProductsModel.DoesNotExist:
            messages.error(request, 'المنتج غير موجود!' if translation.get_language() == 'ar' else 'Product not found!')
            return redirect('viewProducts')

    if request.method == 'POST':
        form = ProductShipmentForm(request.POST, club=club)
        if form.is_valid():
            shipment = form.save()
            messages.success(request, 'تمت إضافة الشحنة بنجاح!' if translation.get_language() == 'ar' else 'Shipment added successfully!')
            return redirect('view_product_shipments', product_id=shipment.product.id)
        else:
            messages.error(request, 'يرجى تصحيح الأخطاء أدناه.' if translation.get_language() == 'ar' else 'Please correct the errors below.')
    else:
        form = ProductShipmentForm(club=club)
        # Pre-select the product if provided
        if product:
            form.fields['product'].initial = product

    context = {
        'form': form,
        'product': product,
        'club': club,
        'LANGUAGE_CODE': translation.get_language(),
        'is_edit': False,
    }

    return render(request, 'club_dashboard/products/ProductsStock/add_edit_shipment.html', context)

def view_product_shipments(request, product_id):
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)


    try:
        product = ProductsModel.objects.get(id=product_id, club=club)
    except ProductsModel.DoesNotExist:
        messages.error(request, 'المنتج غير موجود!')
        return redirect('viewProducts')

    shipments = ProductShipment.objects.filter(product=product).order_by('-created_at')

    expiring_soon_count = sum(1 for s in shipments if s.is_expiring_soon)
    expired_count = sum(1 for s in shipments if s.is_expired)
    valid_count = len(shipments) - expiring_soon_count - expired_count

    total_quantity = sum(s.quantity for s in shipments)
    expiring_soon_quantity = sum(s.quantity for s in shipments if s.is_expiring_soon)
    expired_quantity = sum(s.quantity for s in shipments if s.is_expired)
    valid_quantity = total_quantity - expiring_soon_quantity - expired_quantity

    context = {
        'product': product,
        'shipments': shipments,
        'stats': {
            'total_count': len(shipments),
            'expiring_soon_count': expiring_soon_count,
            'expired_count': expired_count,
            'valid_count': valid_count,
            'total_quantity': total_quantity,
            'expiring_soon_quantity': expiring_soon_quantity,
            'expired_quantity': expired_quantity,
            'valid_quantity': valid_quantity,
        },
        'club':club,
    }
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/products/ProductsStock/view_product_shipments.html', context)

def product_details(request, product_id):
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    try:
        product = ProductsModel.objects.get(id=product_id, club=club)
    except ProductsModel.DoesNotExist:
        messages.error(request, 'المنتج غير موجود!')
        return redirect('viewProducts')

    product_images = product.product_images.all()

    shipments = ProductShipment.objects.filter(product=product).order_by('-created_at')

    expiring_soon_count = sum(1 for s in shipments if s.is_expiring_soon)
    expired_count = sum(1 for s in shipments if s.is_expired)
    valid_count = len(shipments) - expiring_soon_count - expired_count

    total_quantity = sum(s.quantity for s in shipments)
    expiring_soon_quantity = sum(s.quantity for s in shipments if s.is_expiring_soon)
    expired_quantity = sum(s.quantity for s in shipments if s.is_expired)
    valid_quantity = total_quantity - expiring_soon_quantity - expired_quantity

    context = {
        'product': product,
        'product_images': product_images,
        'shipments': shipments,
        'stats': {
            'total_count': len(shipments),
            'expiring_soon_count': expiring_soon_count,
            'expired_count': expired_count,
            'valid_count': valid_count,
            'total_quantity': total_quantity,
            'expiring_soon_quantity': expiring_soon_quantity,
            'expired_quantity': expired_quantity,
            'valid_quantity': valid_quantity,
        },
        'club':club,
    }
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/products/ProductsStock/product_details.html', context)


@login_required
def club_financial_dashboard(request):
    user = request.user

    if not hasattr(user.userprofile, 'director_profile') or not user.userprofile.director_profile:
        messages.error(request, "Unauthorized access.")
        return redirect('home')

    club = user.userprofile.director_profile.club

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date:
        start_date = (timezone.now() - timezone.timedelta(days=30)).strftime('%Y-%m-%d')
    if not end_date:
        end_date = timezone.now().strftime('%Y-%m-%d')

    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
        end_date_obj = end_date_obj.replace(hour=23, minute=59, second=59)
    except ValueError:
        start_date_obj = timezone.now() - timezone.timedelta(days=30)
        end_date_obj = timezone.now()

    orders = Order.objects.filter(
        club=club,
        created_at__gte=start_date_obj,
        created_at__lte=end_date_obj
    ).order_by('-created_at')

    credit_card_orders = orders.filter(payment_method='credit_card')
    cash_orders = orders.filter(payment_method='cash_on_delivery')

    total_revenue = orders.filter(status__in=['confirmed', 'completed']).aggregate(Sum('total_price'))['total_price__sum'] or 0
    pending_revenue = orders.filter(status='pending').aggregate(Sum('total_price'))['total_price__sum'] or 0

    credit_card_revenue = credit_card_orders.filter(status__in=['confirmed', 'completed']).aggregate(Sum('total_price'))['total_price__sum'] or 0
    cash_revenue = cash_orders.filter(status__in=['confirmed', 'completed']).aggregate(Sum('total_price'))['total_price__sum'] or 0

    total_confirmed_revenue = total_revenue
    credit_card_percentage = 0
    cash_percentage = 0
    if total_confirmed_revenue > 0:
        credit_card_percentage = round((credit_card_revenue / total_confirmed_revenue) * 100)
        cash_percentage = round((cash_revenue / total_confirmed_revenue) * 100)

    months_data = []
    for i in range(6):
        month_date = timezone.now() - timezone.timedelta(days=30 * i)
        month_start = month_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if month_date.month == 12:
            month_end = month_date.replace(year=month_date.year + 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0) - timezone.timedelta(seconds=1)
        else:
            month_end = month_date.replace(month=month_date.month + 1, day=1, hour=0, minute=0, second=0, microsecond=0) - timezone.timedelta(seconds=1)

        month_orders = Order.objects.filter(
            club=club,
            created_at__gte=month_start,
            created_at__lte=month_end,
            status__in=['confirmed', 'completed']
        )

        month_credit = month_orders.filter(payment_method='credit_card').aggregate(Sum('total_price'))['total_price__sum'] or 0
        month_cash = month_orders.filter(payment_method='cash_on_delivery').aggregate(Sum('total_price'))['total_price__sum'] or 0

        months_data.append({
            'month': month_start.strftime('%B %Y'),
            'credit_card': float(month_credit),
            'cash': float(month_cash),
            'total': float(month_credit + month_cash)
        })

    months_data.reverse()

    order_items = OrderItem.objects.filter(order__club=club, order__status__in=['confirmed', 'completed'])

    product_sales = {}
    service_sales = {}

    for item in order_items:
        if item.product:
            product_id = item.product.id
            product_name = item.product.title
            if product_id in product_sales:
                product_sales[product_id]['quantity'] += item.quantity
                product_sales[product_id]['revenue'] += float(item.price * item.quantity)
            else:
                product_sales[product_id] = {
                    'name': product_name,
                    'quantity': item.quantity,
                    'revenue': float(item.price * item.quantity)
                }

        if item.service:
            service_id = item.service.id
            service_name = item.service.title
            if service_id in service_sales:
                service_sales[service_id]['quantity'] += item.quantity
                service_sales[service_id]['revenue'] += float(item.service.effective_price * item.quantity)
            else:
                service_sales[service_id] = {
                    'name': service_name,
                    'quantity': item.quantity,
                    'revenue': float(item.price * item.quantity)
                }

    top_products = sorted([v for k, v in product_sales.items()], key=lambda x: x['revenue'], reverse=True)[:5]
    top_services = sorted([v for k, v in service_sales.items()], key=lambda x: x['revenue'], reverse=True)[:5]

    context = {
        'all_orders': orders,
        'orders': orders,
        'credit_card_orders': credit_card_orders,
        'cash_orders': cash_orders,
        'total_revenue': total_revenue,
        'pending_revenue': pending_revenue,
        'credit_card_revenue': credit_card_revenue,
        'cash_revenue': cash_revenue,
        'credit_card_percentage': credit_card_percentage,
        'cash_percentage': cash_percentage,
        'start_date': start_date,
        'end_date': end_date,
        'months_data': json.dumps(months_data),
        'top_products': top_products,
        'top_services': top_services,
    }
    context['LANGUAGE_CODE'] = translation.get_language()
    return render(request, 'club_dashboard/financial/dashboard.html', context)

@login_required
def view_director_profile(request):
    """View the director's profile"""
    try:
        # Get the current user's UserProfile
        userprofile = request.user.userprofile

        # Check if user is a director
        if not userprofile.director_profile:
            return HttpResponseForbidden("You don't have permission to view this page")

        director = userprofile.director_profile

        context = {
            'director': director,
            'userprofile': userprofile,
        }
        context['LANGUAGE_CODE'] = translation.get_language()
        return render(request, 'accounts/profiles/Director/ViewDirectorProfile.html', context)
    except UserProfile.DoesNotExist:
        return HttpResponseForbidden("User profile not found")

@login_required
def edit_director_profile(request):
    """Edit the director's profile"""
    try:
        # Get the current user's UserProfile
        userprofile = request.user.userprofile

        # Check if user is a director
        if not userprofile.director_profile:
            return HttpResponseForbidden("You don't have permission to edit this page")

        director = userprofile.director_profile

        if request.method == 'POST':
            form = DirectorProfileForm(request.POST, request.FILES, instance=director)
            if form.is_valid():
                # Save director profile form
                director = form.save()

                # Handle profile image upload
                if 'profile_image_base64' in request.FILES:
                    image_file = request.FILES['profile_image_base64']
                    # Convert the image to base64
                    encoded_image = base64.b64encode(image_file.read()).decode('utf-8')
                    image_data = f"data:image/{image_file.content_type.split('/')[-1]};base64,{encoded_image}"

                    # Update the user profile
                    userprofile.profile_image_base64 = image_data
                    userprofile.save()

                return redirect(reverse('view_director_profile'))
        else:
            form = DirectorProfileForm(instance=director)

        context = {
            'form': form,
            'director': director,
            'userprofile': userprofile,
        }
        context['LANGUAGE_CODE'] = translation.get_language()
        return render(request, 'accounts/settings/Director/EditDirectorProfile.html', context)
    except UserProfile.DoesNotExist:
        return HttpResponseForbidden("User profile not found")

def handle_uploaded_image(image_file):
    """Convert uploaded image to base64 string"""
    if not image_file:
        return None

    # Check file size (limit to 2MB)
    if image_file.size > 2 * 1024 * 1024:
        raise ValueError("Image file too large (max 2MB)")

    # Get file extension
    file_extension = image_file.name.split('.')[-1].lower()
    if file_extension not in ['jpg', 'jpeg', 'png', 'gif']:
        raise ValueError("Unsupported file format")

    # Convert to base64
    encoded_image = base64.b64encode(image_file.read()).decode('utf-8')

    # Create data URL based on file type
    if file_extension in ['jpg', 'jpeg']:
        return f"data:image/jpeg;base64,{encoded_image}"
    elif file_extension == 'png':
        return f"data:image/png;base64,{encoded_image}"
    elif file_extension == 'gif':
        return f"data:image/gif;base64,{encoded_image}"

def toggle_dashboard_counts(request):
    """Toggle dashboard counts visibility - accessible only to staff"""
    if request.method == 'POST':
        settings = DashboardSettings.get_settings()
        settings.show_employee_client_counts = not settings.show_employee_client_counts
        settings.save()

        status = "shown" if settings.show_employee_client_counts else "hidden"

        if request.headers.get('Accept') == 'application/json':
            return JsonResponse({
                'success': True,
                'status': settings.show_employee_client_counts,
                'message': f"Dashboard count cards are now {status}!"
            })
        else:
            messages.success(request, f"Dashboard count cards are now {status}!")
            return redirect('/admin/')

    return redirect('/admin/')

def update_club_descriptions(request, club_id):
    club = get_object_or_404(ClubsModel, pk=club_id)
    if request.method == 'POST':
        club.productsDescription = request.POST.get('productsDescription', '')
        club.articlesDescription = request.POST.get('articlesDescription', '')
        club.save()
        messages.success(request, "Descriptions updated successfully.")
    return redirect('club_dashboard_index')

# def update_club_pricing(request, club_id):
#     club = get_object_or_404(ClubsModel, pk=club_id)
#     if request.method == 'POST':
#         new_pricing = []
#         i = 1
#         while f'name_{i}' in request.POST:
#             name = request.POST.get(f'name_{i}')
#             price = request.POST.get(f'price_{i}')
#             features = request.POST.get(f'features_{i}', '')
#             features_list = [f.strip() for f in features.split(',') if f.strip()]
#             new_pricing.append({
#                 'name': name,
#                 'price': price,
#                 'features': features_list
#             })
#             i += 1
#         club.pricing = new_pricing
#         club.save()
#         messages.success(request, "تم تحديث التسعير بنجاح." if request.LANGUAGE_CODE == 'ar' else "Pricing updated successfully.")
#     return redirect('club_dashboard_index')

def vendor_status(request, vendor_id):
    """Show vendor application status"""
    vendor = get_object_or_404(CoachProfile, id=vendor_id)
    return render(request, 'accounts/vendor_status.html', {'vendor': vendor})

from django.contrib.auth.decorators import login_required
@login_required
def vendor_approval_list(request):
    """List of vendors pending approval - only for directors"""
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.account_type != '2':  # Not a director
            messages.error(request, "غير مسموح لك بالوصول إلى هذه الصفحة.")
            return redirect('dashboard')

        # Get director's club
        director_profile = user_profile.director_profile
        if not director_profile or not director_profile.club:
            messages.error(request, "لم يتم العثور على نادي مرتبط بحسابك.")
            return redirect('dashboard')

        # Get pending vendors for this club
        pending_vendors = CoachProfile.objects.filter(
            club=director_profile.club,
            approval_status='pending'
        ).order_by('-created_at')

        return render(request, 'accounts/vendor_approval_list.html', {
            'pending_vendors': pending_vendors,
            'club': director_profile.club
        })

    except UserProfile.DoesNotExist:
        messages.error(request, "ملف المستخدم غير موجود.")
        return redirect('signin')


from accounts.forms import VendorApprovalForm
@login_required
def vendor_approval_detail(request, vendor_id):
    """Approve or reject a vendor application"""
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.account_type != '2':  # Not a director
            messages.error(request, "غير مسموح لك بالوصول إلى هذه الصفحة.")
            return redirect('dashboard')

        director_profile = user_profile.director_profile
        vendor = get_object_or_404(CoachProfile, id=vendor_id, club=director_profile.club)

        if request.method == 'POST':
            form = VendorApprovalForm(request.POST)
            if form.is_valid():
                action = form.cleaned_data['action']
                notes = form.cleaned_data['notes']

                if action == 'approve':
                    vendor.approve(request.user, notes)
                    messages.success(request, f"تم قبول طلب {vendor.full_name} بنجاح!")

                    # Send approval email to vendor
                    send_vendor_approval_email(vendor, approved=True)

                elif action == 'reject':
                    vendor.reject(request.user, notes)
                    messages.success(request, f"تم رفض طلب {vendor.full_name}.")

                    # Send rejection email to vendor
                    send_vendor_approval_email(vendor, approved=False)

                return redirect('vendor_approval_list')
        else:
            form = VendorApprovalForm()

        return render(request, 'accounts/vendor_approval_detail.html', {
            'vendor': vendor,
            'form': form
        })

    except UserProfile.DoesNotExist:
        messages.error(request, "ملف المستخدم غير موجود.")
        return redirect('signin')

from django.core.mail import send_mail
from django.conf import settings
def send_vendor_approval_email(vendor, approved=True):
    """Send email to vendor about approval/rejection"""
    try:
        if approved:
            subject = f"تم قبول طلبك - {vendor.business_name}"
            message = f"""
            مرحباً {vendor.full_name},
            
            تم قبول طلب تسجيلك كبائع في نادي {vendor.club.name}.
            
            سيتم إنشاء حسابك قريباً وسيتم إرسال بيانات الدخول إليك.
            
            شكراً لانضمامك إلينا!
            """
        else:
            subject = f"طلب التسجيل - {vendor.business_name}"
            message = f"""
            مرحباً {vendor.full_name},
            
            نأسف لإبلاغك أنه لم يتم قبول طلب تسجيلك كبائع في نادي {vendor.club.name}.
            
            {vendor.approval_notes if vendor.approval_notes else ''}
            
            شكراً لك على اهتمامك.
            """

        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [vendor.email],
            fail_silently=True,
        )
    except Exception as e:
        print(f"Error sending approval email: {e}")

@login_required
def vendor_approval_action(request, vendor_id):
    """Handle vendor approval/rejection actions"""
    try:
        vendor = CoachProfile.objects.get(id=vendor_id)
        user_profile = UserProfile.objects.get(user=request.user)

        if user_profile.account_type != '2':  # Only directors can approve
            messages.error(request, "You don't have permission to perform this action.")
            return redirect('vendor_approval_list')

        if request.method == 'POST':
            action = request.POST.get('action')
            notes = request.POST.get('notes', '')

            if action == 'approve':
                vendor.approve(request.user, notes)
                messages.success(request, f"Vendor {vendor.full_name} approved successfully.")
            elif action == 'reject':
                vendor.reject(request.user, notes)
                messages.success(request, f"Vendor {vendor.full_name} rejected.")
            else:
                messages.error(request, "Invalid action.")

            return redirect('vendor_approval_list')

    except CoachProfile.DoesNotExist:
        messages.error(request, "Vendor not found.")
        return redirect('vendor_approval_list')
    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('vendor_approval_list')

from django.views.generic import DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from accounts.models import CoachProfile

class VendorApprovalDetailView(LoginRequiredMixin, DetailView):
    model = CoachProfile
    template_name = 'club_dashboard/vendor_approval_detail.html'
    context_object_name = 'vendor'

    def get_queryset(self):
        # Only show pending vendors
        return CoachProfile.objects.filter(approval_status='pending')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add any additional context you need
        return context



from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db.models import Count, Q
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import timedelta
from .models import Category, SubCategory
from .forms import CategoryForm, SubCategoryForm


@login_required
def category_list(request):
    """View to display all categories and subcategories"""
    # Get search query
    search_query = request.GET.get('search', '')

    # Filter categories based on search
    categories = Category.objects.all().order_by('name')
    if search_query:
        categories = categories.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    # Annotate with subcategory count
    categories = categories.annotate(subcategory_count=Count('subcategories'))

    # Get subcategories
    subcategories = SubCategory.objects.select_related('category').order_by('name')
    if search_query:
        subcategories = subcategories.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(category__name__icontains=search_query)
        )

    # Pagination for categories
    category_paginator = Paginator(categories, 10)
    category_page = request.GET.get('category_page')
    categories = category_paginator.get_page(category_page)

    # Pagination for subcategories
    subcategory_paginator = Paginator(subcategories, 10)
    subcategory_page = request.GET.get('subcategory_page')
    subcategories = subcategory_paginator.get_page(subcategory_page)

    # Statistics
    stats = {
        'total_categories': Category.objects.count(),
        'active_categories': Category.objects.filter(is_active=True).count(),
        'total_subcategories': SubCategory.objects.count(),
        'active_subcategories': SubCategory.objects.filter(is_active=True).count(),
    }

    # Recent activities (last 30 days)
    recent_date = timezone.now() - timedelta(days=30)
    recent_categories = Category.objects.filter(created_at__gte=recent_date).order_by('-created_at')[:5]
    recent_subcategories = SubCategory.objects.filter(created_at__gte=recent_date).order_by('-created_at')[:5]

    context = {
        'categories': categories,
        'subcategories': subcategories,
        'stats': stats,
        'recent_categories': recent_categories,
        'recent_subcategories': recent_subcategories,
        'search_query': search_query,
    }

    return render(request, 'club_dashboard/categories/category_list.html', context)


@login_required
def add_category(request):
    """View to add a new category"""
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Category "{category.name}" has been created successfully!')
            return redirect('category_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CategoryForm()

    context = {
        'form': form,
        'title': 'Add New Category',
        'submit_text': 'Create Category',
    }
    return render(request, 'club_dashboard/categories/category_form.html', context)


@login_required
def edit_category(request, category_id):
    """View to edit an existing category"""
    category = get_object_or_404(Category, id=category_id)

    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES, instance=category)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Category "{category.name}" has been updated successfully!')
            return redirect('category_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CategoryForm(instance=category)

    context = {
        'form': form,
        'category': category,
        'title': f'Edit Category - {category.name}',
        'submit_text': 'Update Category',
    }
    return render(request, 'club_dashboard/categories/category_form.html', context)


@login_required
def delete_category(request, category_id):
    """View to delete a category"""
    category = get_object_or_404(Category, id=category_id)

    # Check if category has subcategories
    if category.subcategories.exists():
        messages.error(request, f'Cannot delete category "{category.name}" because it has subcategories. Please delete or move the subcategories first.')
        return redirect('category_list')

    if request.method == 'POST':
        category_name = category.name
        category.delete()
        messages.success(request, f'Category "{category_name}" has been deleted successfully!')
        return redirect('category_list')

    context = {
        'category': category,
        'subcategory_count': category.subcategories.count(),
    }
    return render(request, 'club_dashboard/categories/category_delete.html', context)


@login_required
def add_subcategory(request):
    """View to add a new subcategory"""
    if request.method == 'POST':
        form = SubCategoryForm(request.POST, request.FILES)
        if form.is_valid():
            subcategory = form.save()
            messages.success(request, f'Subcategory "{subcategory.name}" has been created successfully!')
            return redirect('category_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = SubCategoryForm()

    context = {
        'form': form,
        'title': 'Add New Subcategory',
        'submit_text': 'Create Subcategory',
    }
    return render(request, 'club_dashboard/categories/subcategory_form.html', context)


@login_required
def edit_subcategory(request, subcategory_id):
    """View to edit an existing subcategory"""
    subcategory = get_object_or_404(SubCategory, id=subcategory_id)

    if request.method == 'POST':
        form = SubCategoryForm(request.POST, request.FILES, instance=subcategory)
        if form.is_valid():
            subcategory = form.save()
            messages.success(request, f'Subcategory "{subcategory.name}" has been updated successfully!')
            return redirect('category_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = SubCategoryForm(instance=subcategory)

    context = {
        'form': form,
        'subcategory': subcategory,
        'title': f'Edit Subcategory - {subcategory.name}',
        'submit_text': 'Update Subcategory',
    }
    return render(request, 'club_dashboard/categories/subcategory_form.html', context)


@login_required
def delete_subcategory(request, subcategory_id):
    """View to delete a subcategory"""
    subcategory = get_object_or_404(SubCategory, id=subcategory_id)

    if request.method == 'POST':
        subcategory_name = subcategory.name
        category_name = subcategory.category.name
        subcategory.delete()
        messages.success(request, f'Subcategory "{subcategory_name}" from "{category_name}" has been deleted successfully!')
        return redirect('category_list')

    context = {
        'subcategory': subcategory,
    }
    return render(request, 'club_dashboard/categories/subcategory_delete.html', context)


@login_required
def category_detail(request, category_id):
    """View to show category details with its subcategories"""
    category = get_object_or_404(Category, id=category_id)
    subcategories = category.subcategories.all().order_by('name')

    # Pagination for subcategories
    paginator = Paginator(subcategories, 12)
    page = request.GET.get('page')
    subcategories = paginator.get_page(page)

    context = {
        'category': category,
        'subcategories': subcategories,
    }
    return render(request, 'club_dashboard/categories/category_detail.html', context)


# AJAX Views for dynamic functionality
@login_required
def get_subcategories(request, category_id):
    """AJAX view to get subcategories for a specific category"""
    category = get_object_or_404(Category, id=category_id)
    subcategories = category.subcategories.filter(is_active=True).values('id', 'name')
    return JsonResponse({'subcategories': list(subcategories)})


@login_required
def toggle_category_status(request, category_id):
    """AJAX view to toggle category active status"""
    if request.method == 'POST':
        category = get_object_or_404(Category, id=category_id)
        category.is_active = not category.is_active
        category.save()

        status_text = 'activated' if category.is_active else 'deactivated'
        messages.success(request, f'Category "{category.name}" has been {status_text}!')

        return JsonResponse({
            'success': True,
            'is_active': category.is_active,
            'message': f'Category {status_text} successfully!'
        })

    return JsonResponse({'success': False, 'message': 'Invalid request method.'})


@login_required
def toggle_subcategory_status(request, subcategory_id):
    """AJAX view to toggle subcategory active status"""
    if request.method == 'POST':
        subcategory = get_object_or_404(SubCategory, id=subcategory_id)
        subcategory.is_active = not subcategory.is_active
        subcategory.save()

        status_text = 'activated' if subcategory.is_active else 'deactivated'
        messages.success(request, f'Subcategory "{subcategory.name}" has been {status_text}!')

        return JsonResponse({
            'success': True,
            'is_active': subcategory.is_active,
            'message': f'Subcategory {status_text} successfully!'
        })

    return JsonResponse({'success': False, 'message': 'Invalid request method.'})

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.utils import timezone, translation
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
from .models import ProductsModel, CoachProfile, ClubsModel, ProductImg
from .forms import ProductApprovalForm

@login_required
def manage_products(request):
    """View to manage all products with approval status"""
    context = {}
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)


    # Get filter parameters
    status_filter = request.GET.get('status', 'all')
    search_query = request.GET.get('search', '')

    # Base queryset for products in this club
    products = ProductsModel.objects.filter(club=club).select_related(
        'creator', 'creator__userprofile', 'creator__userprofile__Coach_profile'
    ).prefetch_related('product_images', 'classification')

    # Apply status filter
    if status_filter == 'pending':
        products = products.filter(approval_status='pending')
    elif status_filter == 'approved':
        products = products.filter(approval_status='approved')
    elif status_filter == 'rejected':
        products = products.filter(approval_status='rejected')

    # Apply search filter
    if search_query:
        products = products.filter(
            Q(title__icontains=search_query) |
            Q(creator__userprofile__Coach_profile__business_name__icontains=search_query) |
            Q(creator__userprofile__Coach_profile__full_name__icontains=search_query) |
            Q(creator__email__icontains=search_query)
        )

    # Order by creation date (newest first)
    products = products.order_by('-creation_date')

    # Get statistics
    stats = {
        'total': ProductsModel.objects.filter(club=club).count(),
        'pending': ProductsModel.objects.filter(club=club, approval_status='pending').count(),
        'approved': ProductsModel.objects.filter(club=club, approval_status='approved').count(),
        'rejected': ProductsModel.objects.filter(club=club, approval_status='rejected').count(),
    }

    # Get latest pending products for quick review
    latest_pending = ProductsModel.objects.filter(
        club=club,
        approval_status='pending'
    ).select_related(
        'creator', 'creator__userprofile', 'creator__userprofile__Coach_profile'
    ).order_by('-creation_date')[:3]

    # Pagination
    paginator = Paginator(products, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context.update({
        'products': page_obj,
        'stats': stats,
        'latest_pending': latest_pending,
        'status_filter': status_filter,
        'search_query': search_query,
        'club': club,
        'LANGUAGE_CODE': translation.get_language()
    })

    return render(request, 'club_dashboard/products/manage_products.html', context)

@login_required
def approve_product(request, product_id):
    """Approve a product"""
    product = get_object_or_404(ProductsModel, id=product_id)
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    # Check permissions
    if product.club != club:
        messages.error(request, 'غير مسموح لك بتعديل هذا المنتج')
        return redirect('manage_products')

    if request.method == 'POST':
        notes = request.POST.get('approval_notes', '')

        # Update product status
        product.approval_status = 'approved'
        product.approved_at = timezone.now()
        product.approved_by = user
        product.approval_notes = notes
        product.is_enabled = True
        product.save()

        # Send email notification to vendor
        try:
            vendor_profile = product.creator.userprofile.Coach_profile
            subject = f"تم قبول منتجك: {product.title}"
            context = {
                'product': product,
                'vendor': vendor_profile,
                'notes': notes,
                'club': club
            }
            html_message = render_to_string('emails/product_approved.html', context)
            plain_message = strip_tags(html_message)

            send_mail(
                subject=subject,
                message=plain_message,
                html_message=html_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[product.creator.email],
                fail_silently=True,
            )
        except Exception as e:
            print(f"Error sending email: {e}")

        messages.success(request, f'تم قبول المنتج "{product.title}" بنجاح')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'تم قبول المنتج بنجاح'})

    return redirect('manage_products')

@login_required
def reject_product(request, product_id):
    """Reject a product"""
    product = get_object_or_404(ProductsModel, id=product_id)
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    # Check permissions
    if product.club != club:
        messages.error(request, 'غير مسموح لك بتعديل هذا المنتج')
        return redirect('manage_products')

    if request.method == 'POST':
        rejection_reason = request.POST.get('rejection_reason', '')

        # Update product status
        product.approval_status = 'rejected'
        product.approved_at = timezone.now()
        product.approved_by = user
        product.approval_notes = rejection_reason
        product.is_enabled = False
        product.save()

        # Send email notification to vendor
        try:
            vendor_profile = product.creator.userprofile.Coach_profile
            subject = f"تم رفض منتجك: {product.title}"
            context = {
                'product': product,
                'vendor': vendor_profile,
                'rejection_reason': rejection_reason,
                'club': club
            }
            html_message = render_to_string('products/emails/product_rejected.html', context)
            plain_message = strip_tags(html_message)

            send_mail(
                subject=subject,
                message=plain_message,
                html_message=html_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[product.creator.email],
                fail_silently=True,
            )
        except Exception as e:
            print(f"Error sending email: {e}")

        messages.success(request, f'تم رفض المنتج "{product.title}"')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'تم رفض المنتج'})

    return redirect('manage_products')

@login_required
def product_detail(request, product_id):
    """View detailed information about a product"""
    product = get_object_or_404(ProductsModel, id=product_id)
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    # Check permissions
    if product.club != club:
        messages.error(request, 'غير مسموح لك بعرض هذا المنتج')
        return redirect('manage_products')

    # Get product images
    product_images = product.product_images.all()

    # Get number of orders for this product
    order_count = OrderItem.objects.filter(product=product).count()

    context = {
        'product': product,
        'product_images': product_images,
        'vendor_profile': product.creator.userprofile.Coach_profile if hasattr(product.creator, 'userprofile') else None,
        'club': club,
        'LANGUAGE_CODE': translation.get_language(),
        'order_count': order_count
    }

    return render(request, 'club_dashboard/products/product_detail.html', context)

@login_required
def bulk_approve_products(request):
    """Bulk approve multiple products"""
    if request.method == 'POST':
        product_ids = request.POST.getlist('product_ids')
        user = request.user
        club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

        if not product_ids:
            messages.error(request, 'لم يتم تحديد أي منتجات')
            return redirect('manage_products')

        # Update products
        updated_count = ProductsModel.objects.filter(
            id__in=product_ids,
            club=club,
            approval_status='pending'
        ).update(
            approval_status='approved',
            approved_at=timezone.now(),
            approved_by=user,
            is_enabled=True
        )

        messages.success(request, f'تم قبول {updated_count} منتج بنجاح')

    return redirect('manage_products')

@login_required
def bulk_reject_products(request):
    """Bulk reject multiple products"""
    if request.method == 'POST':
        product_ids = request.POST.getlist('product_ids')
        rejection_reason = request.POST.get('bulk_rejection_reason', 'تم الرفض بواسطة الإدارة')
        user = request.user
        club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

        if not product_ids:
            messages.error(request, 'لم يتم تحديد أي منتجات')
            return redirect('manage_products')

        # Update products
        updated_count = ProductsModel.objects.filter(
            id__in=product_ids,
            club=club,
            approval_status='pending'
        ).update(
            approval_status='rejected',
            approved_at=timezone.now(),
            approved_by=user,
            approval_notes=rejection_reason,
            is_enabled=False
        )

        messages.success(request, f'تم رفض {updated_count} منتج')

    return redirect('manage_products')



# club_dashboard/views.py

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils import timezone
from .models import Commission, VendorCommissionAssignment
from .forms import CommissionForm
from accounts.models import CoachProfile

from accounts.models import CoachProfile
@login_required
def commission_list(request):
    """List all commissions with filtering and pagination"""
    # Get filter parameters
    commission_type = request.GET.get('type', '')
    classification = request.GET.get('classification', '')
    status = request.GET.get('status', '')
    search = request.GET.get('search', '')
    user = request.user

    # Get the club
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    # Base queryset
    commissions = Commission.objects.filter(club=club)

    # Apply filters
    if commission_type:
        commissions = commissions.filter(commission_type=commission_type)

    if classification:
        commissions = commissions.filter(vendor_classification=classification)

    if status == 'active':
        commissions = commissions.filter(is_active=True)
    elif status == 'inactive':
        commissions = commissions.filter(is_active=False)

    if search:
        commissions = commissions.filter(
            Q(name__icontains=search) |
            Q(commission_rate__icontains=search)
        )


    latest_vendor_commissions = Commission.objects.filter(
        club=club,
        commission_type='vendor'
    ).order_by('-created_at')[:4]

    # Pagination
    paginator = Paginator(commissions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get unique vendor classifications for the club
    vendor_classifications = Commission.objects.filter(
        club=club,
        commission_type='vendor'
    ).values_list('vendor_classification', flat=True).distinct()

    # Create choices list for the filter form
    classification_choices = [(c, c.capitalize()) for c in vendor_classifications]


    # Statistics
    stats = {
        'total_classifications': Commission.objects.filter(club=club).values('vendor_classification').distinct().count(),
        'total_offers': Commission.objects.filter(club=club,commission_type='time_period').distinct().count(),  # Assuming you have an Offer model
        'total_vendors': CoachProfile.objects.filter(club=club).count(),  # Assuming you have a Vendor model
        'active_commissions': Commission.objects.filter(club=club, is_active=True).count(),
    }

    context = {
        'page_obj': page_obj,
        'stats': stats,
        'current_filters': {
            'type': commission_type,
            'classification': classification,
            'status': status,
            'search': search,
        },
        'commission_types': Commission.COMMISSION_TYPES,
        'vendor_classifications': classification_choices,  # Updated to use dynamic classifications
        'latest_vendor_commissions': latest_vendor_commissions,
    }

    return render(request, 'club_dashboard/commissions/list.html', context)

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages

@login_required
def commission_create(request):
    """Create a new commission"""

    print("DEBUG: commission_create view called")

    user = request.user
    print(f"DEBUG: User: {user}, Authenticated: {user.is_authenticated}")

    # Try getting the club
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)
    print(f"DEBUG: Club determined from user profile: {club}")

    if request.method == 'POST':
        print("DEBUG: Request method is POST")
        form = CommissionForm(request.POST, club=club)
        print(f"DEBUG: Form initialized with POST data and club: {form}")

        if form.is_valid():
            print("DEBUG: Form is valid")
            commission = form.save(commit=False)
            commission.club = club
            commission.created_by = user
            commission.save()
            print(f"DEBUG: Commission created and saved: {commission}")

            messages.success(request, 'تم إنشاء العمولة بنجاح')
            return redirect('commission_list')
        else:
            print("DEBUG: Form is not valid")
            print(f"DEBUG: Form errors: {form.errors}")

    else:
        print("DEBUG: Request method is GET")
        form = CommissionForm(club=club)
        print("DEBUG: Form initialized with empty data and club")

    return render(request, 'club_dashboard/commissions/create.html', {'form': form})


@login_required
def commission_edit(request, commission_id):
    """Edit an existing commission"""
    user = request.user
    commission = get_object_or_404(Commission, id=commission_id, club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None))

    if request.method == 'POST':
        form = CommissionForm(request.POST, instance=commission, club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None))
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تحديث العمولة بنجاح')
            return redirect('commission_list')
    else:
        form = CommissionForm(instance=commission, club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None))

    return render(request, 'club_dashboard/commissions/edit.html', {
        'form': form,
        'commission': commission
    })

@login_required
@require_http_methods(["POST"])
def commission_delete(request, commission_id):
    """Delete a commission"""
    user = request.user
    commission = get_object_or_404(Commission, id=commission_id, club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None))

    # Check if commission is assigned to any vendors
    assigned_vendors_count = VendorCommissionAssignment.objects.filter(commission=commission).count()

    if assigned_vendors_count > 0:
        messages.error(request, f'لا يمكن حذف هذه العمولة لأنها مخصصة لـ {assigned_vendors_count} بائع')
    else:
        commission_name = commission.name
        commission.delete()
        messages.success(request, f'تم حذف العمولة "{commission_name}" بنجاح')

    return redirect('club_dashboard:commission_list')

@login_required
@require_http_methods(["POST"])
def commission_toggle_status(request, commission_id):
    """Toggle commission active/inactive status"""
    user = request.user
    commission = get_object_or_404(Commission, id=commission_id, club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None))

    commission.is_active = not commission.is_active
    commission.save()

    status_text = 'مفعلة' if commission.is_active else 'معطلة'
    messages.success(request, f'تم تغيير حالة العمولة إلى {status_text}')

    return JsonResponse({
        'success': True,
        'is_active': commission.is_active,
        'status_text': status_text
    })

@login_required
def commission_detail(request, commission_id):
    """View commission details and assigned vendors"""
    user = request.user
    commission = get_object_or_404(Commission, id=commission_id, club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None))

    # Get assigned vendors for vendor type commissions
    assigned_vendors = []
    if commission.commission_type == 'vendor':
        assigned_vendors = VendorCommissionAssignment.objects.filter(
            commission=commission
        ).select_related('vendor')

    # Get active time period commissions (for reference)
    active_time_commissions = Commission.objects.filter(
        club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None),
        commission_type='time_period',
        is_active=True,
        start_date__lte=timezone.now().date(),
        end_date__gte=timezone.now().date()
    )

    context = {
        'commission': commission,
        'assigned_vendors': assigned_vendors,
        'active_time_commissions': active_time_commissions,
    }

    return render(request, 'club_dashboard/commissions/detail.html', context)

@login_required
def vendor_commission_management(request):
    """Manage vendor commission assignments"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    # Get all approved vendors for the club
    vendors = CoachProfile.objects.filter(
        club=club,
        approval_status='approved'
    ).select_related('commission_assignment__commission')

    # Get available vendor commissions
    available_commissions = Commission.objects.filter(
        club=club,
        commission_type='vendor',
        is_active=True
    )

    # Get unique vendor classifications for the club
    vendor_classifications = Commission.objects.filter(
        club=club,
        commission_type='vendor'
    ).values_list('vendor_classification', flat=True).distinct()

    # Create stats for each classification
    classification_stats = []
    for classification in vendor_classifications:
        count = CoachProfile.objects.filter(
            club=club,
            approval_status='approved',
            vendor_classification=classification
        ).count()
        classification_stats.append((classification, classification.capitalize(), count))

    # Handle bulk assignment
    if request.method == 'POST':
        vendor_id = request.POST.get('vendor_id')
        commission_id = request.POST.get('commission_id')

        if vendor_id and commission_id:
            vendor = get_object_or_404(CoachProfile, id=vendor_id, club=club)
            commission = get_object_or_404(Commission, id=commission_id, club=club)

            # Update or create assignment
            assignment, created = VendorCommissionAssignment.objects.get_or_create(
                vendor=vendor,
                defaults={'commission': commission}
            )

            if not created:
                assignment.commission = commission
                assignment.save()

            # Update vendor classification
            vendor.vendor_classification = commission.vendor_classification
            vendor.save()

            action = 'تم تخصيص' if created else 'تم تحديث'
            messages.success(request, f'{action} العمولة للبائع {vendor.full_name} بنجاح')

    # Pagination
    paginator = Paginator(vendors, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'available_commissions': available_commissions,
        'vendor_classifications': classification_stats,  # Updated to use dynamic stats
    }

    return render(request, 'club_dashboard/commissions/vendor_management.html', context)

@login_required
def commission_analytics(request):
    """Commission analytics and reports"""
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    # Get unique vendor classifications for the club
    vendor_classifications = Commission.objects.filter(
        club=club,
        commission_type='vendor'
    ).values_list('vendor_classification', flat=True).distinct()

    # Vendor commission analytics
    vendor_commission_data = []
    for classification in vendor_classifications:
        commission = Commission.objects.filter(
            club=club,
            commission_type='vendor',
            vendor_classification=classification,
            is_active=True
        ).first()

        vendor_count = CoachProfile.objects.filter(
            club=club,
            vendor_classification=classification,
            approval_status='approved'
        ).count()

        vendor_commission_data.append({
            'classification': classification.capitalize(),
            'commission_rate': commission.commission_rate if commission else 0,
            'vendor_count': vendor_count,
            'total_potential_commission': (commission.commission_rate * vendor_count) if commission else 0
        })

    # Time period commissions
    time_period_commissions = Commission.objects.filter(
        club=club,
        commission_type='time_period',
        is_active=True
    ).order_by('start_date')

    # Current active time commission
    current_time_commission = Commission.get_time_period_commission(club)

    context = {
        'vendor_commission_data': vendor_commission_data,
        'time_period_commissions': time_period_commissions,
        'current_time_commission': current_time_commission,
        'total_vendors': CoachProfile.objects.filter(club=club, approval_status='approved').count(),
    }

    return render(request, 'club_dashboard/commissions/analytics.html', context)



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.utils import timezone, translation
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
from accounts.models import  CoachProfile, ClubsModel
from .forms import ServiceApprovalForm
from students.models import ServicesModel


@login_required
def manage_services(request):
    """View to manage all services with approval status"""
    context = {}
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    if not club:
        messages.error(request, 'غير مسموح لك بالوصول لهذه الصفحة')
        return redirect('club_dashboard')

    # Get filter parameters
    status_filter = request.GET.get('status', 'all')
    search_query = request.GET.get('search', '')

    # Base queryset for services in this club
    services = ServicesModel.objects.filter(club=club).select_related(
        'creator', 'creator__userprofile', 'creator__userprofile__Coach_profile'
    ).prefetch_related('coaches', 'classification')

    # Apply status filter
    if status_filter == 'pending':
        services = services.filter(approval_status='pending')
    elif status_filter == 'approved':
        services = services.filter(approval_status='approved')
    elif status_filter == 'rejected':
        services = services.filter(approval_status='rejected')

    # Apply search filter
    if search_query:
        services = services.filter(
            Q(title__icontains=search_query) |
            Q(desc__icontains=search_query) |
            Q(creator__userprofile__Coach_profile__business_name__icontains=search_query) |
            Q(creator__userprofile__Coach_profile__full_name__icontains=search_query) |
            Q(creator__email__icontains=search_query) |
            Q(coaches__business_name__icontains=search_query) |
            Q(coaches__full_name__icontains=search_query)
        ).distinct()

    # Order by creation date (newest first)
    services = services.order_by('-creation_date')

    # Get statistics
    stats = {
        'total': ServicesModel.objects.filter(club=club).count(),
        'pending': ServicesModel.objects.filter(club=club, approval_status='pending').count(),
        'approved': ServicesModel.objects.filter(club=club, approval_status='approved').count(),
        'rejected': ServicesModel.objects.filter(club=club, approval_status='rejected').count(),
    }

    # Get latest pending services for quick review
    latest_pending = ServicesModel.objects.filter(
        club=club,
        approval_status='pending'
    ).select_related(
        'creator', 'creator__userprofile', 'creator__userprofile__Coach_profile'
    ).order_by('-creation_date')[:3]

    # Pagination
    paginator = Paginator(services, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context.update({
        'services': page_obj,
        'stats': stats,
        'latest_pending': latest_pending,
        'status_filter': status_filter,
        'search_query': search_query,
        'club': club,
        'LANGUAGE_CODE': translation.get_language()
    })

    return render(request, 'club_dashboard/services/manage_services.html', context)

@login_required
def approve_service(request, service_id):
    """Approve a service"""
    service = get_object_or_404(ServicesModel, id=service_id)
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    # Check permissions
    if service.club != club:
        messages.error(request, 'غير مسموح لك بتعديل هذه الخدمة')
        return redirect('manage_services')

    if request.method == 'POST':
        notes = request.POST.get('approval_notes', '')

        # Use the model's approve method
        service.approve(user, notes)

        # Send email notification to creator
        try:
            creator_profile = service.creator.userprofile.Coach_profile
            subject = f"تم قبول خدمتك: {service.title}"
            context = {
                'service': service,
                'creator': creator_profile,
                'notes': notes,
                'club': club
            }
            html_message = render_to_string('services/emails/service_approved.html', context)
            plain_message = strip_tags(html_message)

            send_mail(
                subject=subject,
                message=plain_message,
                html_message=html_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[service.creator.email],
                fail_silently=True,
            )
        except Exception as e:
            print(f"Error sending email: {e}")

        messages.success(request, f'تم قبول الخدمة "{service.title}" بنجاح')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'تم قبول الخدمة بنجاح'})

    return redirect('manage_services')

@login_required
def reject_service(request, service_id):
    """Reject a service"""
    service = get_object_or_404(ServicesModel, id=service_id)
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    # Check permissions
    if service.club != club:
        messages.error(request, 'غير مسموح لك بتعديل هذه الخدمة')
        return redirect('manage_services')

    if request.method == 'POST':
        rejection_reason = request.POST.get('rejection_reason', '')

        # Use the model's reject method
        service.reject(user, rejection_reason)

        # Send email notification to creator
        try:
            creator_profile = service.creator.userprofile.Coach_profile
            subject = f"تم رفض خدمتك: {service.title}"
            context = {
                'service': service,
                'creator': creator_profile,
                'rejection_reason': rejection_reason,
                'club': club
            }
            html_message = render_to_string('services/emails/service_rejected.html', context)
            plain_message = strip_tags(html_message)

            send_mail(
                subject=subject,
                message=plain_message,
                html_message=html_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[service.creator.email],
                fail_silently=True,
            )
        except Exception as e:
            print(f"Error sending email: {e}")

        messages.success(request, f'تم رفض الخدمة "{service.title}"')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'تم رفض الخدمة'})

    return redirect('manage_services')

@login_required
def service_detail(request, service_id):
    """View detailed information about a service"""
    service = get_object_or_404(ServicesModel, id=service_id)
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    # Check permissions
    if service.club != club:
        messages.error(request, 'غير مسموح لك بعرض هذه الخدمة')
        return redirect('manage_services')

    # Get service coaches
    service_coaches = service.coaches.all()

    context = {
        'service': service,
        'service_coaches': service_coaches,
        'creator_profile': service.creator.userprofile.Coach_profile if hasattr(service.creator, 'userprofile') else None,
        'club': club,
        'LANGUAGE_CODE': translation.get_language()
    }

    return render(request, 'club_dashboard/services/service_detail.html', context)

@login_required
def bulk_approve_services(request):
    """Bulk approve multiple services"""
    if request.method == 'POST':
        service_ids = request.POST.getlist('service_ids')
        notes = request.POST.get('bulk_notes', '')
        user = request.user
        club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

        if not service_ids:
            messages.error(request, 'لم يتم تحديد أي خدمات')
            return redirect('manage_services')

        # Get services to update
        services_to_approve = ServicesModel.objects.filter(
            id__in=service_ids,
            club=club,
            approval_status='pending'
        )

        updated_count = 0
        for service in services_to_approve:
            service.approve(user, notes)
            updated_count += 1

        messages.success(request, f'تم قبول {updated_count} خدمة بنجاح')

    return redirect('manage_services')

@login_required
def bulk_reject_services(request):
    """Bulk reject multiple services"""
    if request.method == 'POST':
        service_ids = request.POST.getlist('service_ids')
        rejection_reason = request.POST.get('bulk_rejection_reason', 'تم الرفض بواسطة الإدارة')
        user = request.user
        club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

        if not service_ids:
            messages.error(request, 'لم يتم تحديد أي خدمات')
            return redirect('manage_services')

        # Get services to update
        services_to_reject = ServicesModel.objects.filter(
            id__in=service_ids,
            club=club,
            approval_status='pending'
        )

        updated_count = 0
        for service in services_to_reject:
            service.reject(user, rejection_reason)
            updated_count += 1

        messages.success(request, f'تم رفض {updated_count} خدمة')

    return redirect('manage_services')

@login_required
def toggle_service_status(request, service_id):
    """Toggle service enabled/disabled status"""
    service = get_object_or_404(ServicesModel, id=service_id)
    user = request.user
    club = getattr(user.userprofile.director_profile, 'club', None) or getattr(user.userprofile.administrator_profile, 'club', None)

    # Check permissions
    if service.club != club:
        messages.error(request, 'غير مسموح لك بتعديل هذه الخدمة')
        return redirect('manage_services')

    if request.method == 'POST':
        service.is_enabled = not service.is_enabled
        service.save()

        status_text = 'تم تفعيل' if service.is_enabled else 'تم إلغاء تفعيل'
        messages.success(request, f'{status_text} الخدمة "{service.title}"')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': f'{status_text} الخدمة',
                'is_enabled': service.is_enabled
            })

    return redirect('manage_services')




# views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponseForbidden
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum
from django.utils import timezone
from django.urls import reverse
from decimal import Decimal
import json

from .models import RefundDispute, RefundDisputeAttachment, RefundStatus, RefundType, DisputeType
from students.models import Order  # Assuming Order is in students app
from .forms import RefundDisputeForm, RefundDecisionForm , RefundAttachmentForm


@login_required
def refund_dashboard(request):
    """Main dashboard for refund and dispute management"""
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    dispute_type_filter = request.GET.get('dispute_type', '')
    priority_filter = request.GET.get('priority', '')
    search_query = request.GET.get('search', '')

    # Base queryset
    disputes = RefundDispute.objects.select_related(
        'deal', 'client', 'vendor', 'reviewed_by'
    ).prefetch_related('attachments')

    # Apply filters
    if status_filter:
        disputes = disputes.filter(status=status_filter)

    if dispute_type_filter:
        disputes = disputes.filter(dispute_type=dispute_type_filter)

    if priority_filter:
        disputes = disputes.filter(priority=priority_filter)

    if search_query:
        disputes = disputes.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(deal__id__icontains=search_query) |
            Q(client__email__icontains=search_query) |
            Q(vendor__email__icontains=search_query)
        )

    # Get statistics
    stats = {
        'total_disputes': RefundDispute.objects.count(),
        'pending_disputes': RefundDispute.objects.filter(status=RefundStatus.PENDING).count(),
        'overdue_disputes': RefundDispute.objects.filter(
            status__in=[RefundStatus.PENDING, RefundStatus.INVESTIGATING],
            created_at__lt=timezone.now() - timezone.timedelta(days=7)
        ).count(),
        'total_refund_amount': RefundDispute.objects.filter(
            status=RefundStatus.APPROVED
        ).aggregate(
            total=Sum('approved_refund_amount')
        )['total'] or Decimal('0.00'),
        'avg_resolution_days': 5.2,  # You can calculate this properly
    }

    # Pagination
    paginator = Paginator(disputes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'disputes': page_obj,
        'stats': stats,
        'status_choices': RefundStatus.choices,
        'dispute_type_choices': DisputeType.choices,
        'priority_choices': [('low', 'Low'), ('medium', 'Medium'), ('high', 'High'), ('urgent', 'Urgent')],
        'current_filters': {
            'status': status_filter,
            'dispute_type': dispute_type_filter,
            'priority': priority_filter,
            'search': search_query,
        }
    }

    return render(request, 'club_dashboard/refunds/dashboard.html', context)


@login_required
def refund_detail(request, dispute_id):
    """Detailed view of a specific refund/dispute"""

    dispute = get_object_or_404(
        RefundDispute.objects.select_related(
            'deal', 'client', 'vendor', 'reviewed_by'
        ).prefetch_related('attachments'),
        id=dispute_id
    )

    # Get order details if available
    order_items = []
    if dispute.deal:
        order_items = dispute.deal.items.select_related('product', 'service').all()

    context = {
        'dispute': dispute,
        'order_items': order_items,
        'can_approve': dispute.can_be_approved(),
        'can_reject': dispute.can_be_rejected(),
        'can_resolve': dispute.can_be_resolved(),
    }

    return render(request, 'club_dashboard/refunds/detail.html', context)


@login_required
@require_http_methods(["POST"])
def approve_refund(request, dispute_id):
    """Approve a refund dispute"""

    dispute = get_object_or_404(RefundDispute, id=dispute_id)

    if not dispute.can_be_approved():
        return JsonResponse({
            'success': False,
            'error': 'This dispute cannot be approved in its current state'
        })

    try:
        # Get form data
        approved_amount = Decimal(request.POST.get('approved_amount', '0'))
        vendor_percentage = Decimal(request.POST.get('vendor_percentage', '0'))
        client_percentage = Decimal(request.POST.get('client_percentage', '100'))
        admin_notes = request.POST.get('admin_notes', '')

        # Validate percentages
        if vendor_percentage + client_percentage != 100:
            return JsonResponse({
                'success': False,
                'error': 'Vendor and client percentages must sum to 100%'
            })

        # Validate approved amount
        if approved_amount > dispute.requested_refund_amount:
            return JsonResponse({
                'success': False,
                'error': 'Approved amount cannot exceed requested amount'
            })

        # Update dispute
        dispute.status = RefundStatus.APPROVED
        dispute.approved_refund_amount = approved_amount
        dispute.vendor_percentage = vendor_percentage
        dispute.client_percentage = client_percentage
        dispute.admin_notes = admin_notes
        dispute.reviewed_by = request.user
        dispute.approved_at = timezone.now()
        dispute.save()

        # Process the actual refund (integrate with payment system)
        process_refund_payment(dispute)

        return redirect('detail', dispute_id=dispute.id)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error processing approval: {str(e)}'
        })


@login_required
@require_http_methods(["POST"])
def reject_refund(request, dispute_id):
    """Reject a refund dispute"""


    dispute = get_object_or_404(RefundDispute, id=dispute_id)

    if not dispute.can_be_rejected():
        return JsonResponse({
            'success': False,
            'error': 'This dispute cannot be rejected in its current state'
        })

    try:
        rejection_reason = request.POST.get('rejection_reason', '')
        admin_notes = request.POST.get('admin_notes', '')

        if not rejection_reason:
            return JsonResponse({
                'success': False,
                'error': 'Rejection reason is required'
            })

        dispute.status = RefundStatus.REJECTED
        dispute.rejection_reason = rejection_reason
        dispute.admin_notes = admin_notes
        dispute.reviewed_by = request.user
        dispute.rejected_at = timezone.now()
        dispute.save()

        # Send notification to client
        send_refund_rejection_notification(dispute)

        return redirect('detail', dispute_id=dispute.id)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error processing rejection: {str(e)}'
        })


@login_required
@require_http_methods(["POST"])
def mark_investigating(request, dispute_id):
    """Mark dispute as under investigation"""
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Permission denied'})

    dispute = get_object_or_404(RefundDispute, id=dispute_id)

    if dispute.status != RefundStatus.PENDING:
        return JsonResponse({
            'success': False,
            'error': 'Only pending disputes can be marked as investigating'
        })

    dispute.status = RefundStatus.INVESTIGATING
    dispute.requires_investigation = True
    dispute.reviewed_by = request.user
    dispute.save()

    return redirect('detail', dispute_id=dispute.id)


@login_required
@require_http_methods(["POST"])
def resolve_dispute(request, dispute_id):
    """Mark dispute as resolved"""

    dispute = get_object_or_404(RefundDispute, id=dispute_id)

    if not dispute.can_be_resolved():
        return JsonResponse({
            'success': False,
            'error': 'This dispute cannot be resolved in its current state'
        })

    resolution_notes = request.POST.get('resolution_notes', '')

    dispute.status = RefundStatus.RESOLVED
    dispute.resolution_notes = resolution_notes
    dispute.resolved_at = timezone.now()
    dispute.save()

    return redirect('detail', dispute_id=dispute.id)


@login_required
def bulk_action(request):
    """Handle bulk actions on multiple disputes"""

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})

    try:
        data = json.loads(request.body)
        dispute_ids = data.get('dispute_ids', [])
        action = data.get('action', '')

        if not dispute_ids or not action:
            return JsonResponse({
                'success': False,
                'error': 'Missing dispute IDs or action'
            })

        disputes = RefundDispute.objects.filter(id__in=dispute_ids)

        if action == 'mark_investigating':
            disputes.filter(status=RefundStatus.PENDING).update(
                status=RefundStatus.INVESTIGATING,
                requires_investigation=True,
                reviewed_by=request.user
            )
            message = f'{disputes.count()} disputes marked as investigating'

        elif action == 'assign_priority':
            priority = data.get('priority', 'medium')
            disputes.update(priority=priority)
            message = f'{disputes.count()} disputes priority updated to {priority}'

        else:
            return JsonResponse({
                'success': False,
                'error': 'Invalid action'
            })

        return JsonResponse({
            'success': True,
            'message': message
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error processing bulk action: {str(e)}'
        })


@login_required
def export_disputes(request):
    """Export disputes to CSV"""

    import csv
    from django.http import HttpResponse

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="refund_disputes.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'ID', 'Title', 'Status', 'Dispute Type', 'Client', 'Vendor',
        'Original Amount', 'Requested Amount', 'Approved Amount',
        'Created At', 'Resolved At', 'Priority'
    ])

    disputes = RefundDispute.objects.select_related('client', 'vendor').all()

    for dispute in disputes:
        writer.writerow([
            dispute.id,
            dispute.title,
            dispute.get_status_display(),
            dispute.get_dispute_type_display(),
            dispute.client.username if dispute.client else '',
            dispute.vendor.username if dispute.vendor else '',
            dispute.original_amount,
            dispute.requested_refund_amount,
            dispute.approved_refund_amount or '',
            dispute.created_at.strftime('%Y-%m-%d %H:%M'),
            dispute.resolved_at.strftime('%Y-%m-%d %H:%M') if dispute.resolved_at else '',
            dispute.get_priority_display()
        ])

    return response


# Helper functions
def process_refund_payment(dispute):
    """Process the actual refund payment"""
    # This should integrate with your payment processor
    # For now, just a placeholder
    try:
        # Example: integrate with Stripe, PayPal, etc.
        # payment_processor.refund(
        #     transaction_id=dispute.deal.payment_id,
        #     amount=dispute.approved_refund_amount
        # )

        # Log the refund
        print(f"Processing refund of {dispute.approved_refund_amount} for dispute {dispute.id}")

        # Update order status if needed
        if dispute.deal and dispute.is_full_refund():
            dispute.deal.status = 'refunded'  # Add this status to your Order model
            dispute.deal.save()

    except Exception as e:
        print(f"Error processing refund payment: {e}")
        # You might want to log this error and possibly revert the dispute status


def send_refund_rejection_notification(dispute):
    """Send notification when refund is rejected"""
    # Implement email/SMS notification logic
    try:

        from django.core.mail import send_mail
        from django.conf import settings

        subject = f"Refund Request Rejected - Order #{dispute.deal.id}"
        message = f"""
        Dear {dispute.client.get_full_name()},
        
        Your refund request for Order #{dispute.deal.id} has been rejected.
        
        Reason: {dispute.rejection_reason}
        
        If you have questions, please contact our support team.
        
        Best regards,
        Club Management Team
        """

        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [dispute.client.email],
            fail_silently=True,
        )

    except Exception as e:
        print(f"Error sending rejection notification: {e}")





# API Views for AJAX calls
@login_required
def get_dispute_stats(request):
    """Get dispute statistics for dashboard"""

    stats = {
        'pending': RefundDispute.objects.filter(status=RefundStatus.PENDING).count(),
        'investigating': RefundDispute.objects.filter(status=RefundStatus.INVESTIGATING).count(),
        'approved': RefundDispute.objects.filter(status=RefundStatus.APPROVED).count(),
        'rejected': RefundDispute.objects.filter(status=RefundStatus.REJECTED).count(),
        'resolved': RefundDispute.objects.filter(status=RefundStatus.RESOLVED).count(),
        'overdue': RefundDispute.objects.filter(
            status__in=[RefundStatus.PENDING, RefundStatus.INVESTIGATING],
            created_at__lt=timezone.now() - timezone.timedelta(days=7)
        ).count(),
    }

    return JsonResponse(stats)


@login_required
def update_dispute_priority(request, dispute_id):
    """Update dispute priority"""

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})

    dispute = get_object_or_404(RefundDispute, id=dispute_id)
    priority = request.POST.get('priority')

    if priority not in ['low', 'medium', 'high', 'urgent']:
        return JsonResponse({'success': False, 'error': 'Invalid priority'})

    dispute.priority = priority
    dispute.save()

    return JsonResponse({
        'success': True,
        'message': f'Priority updated to {priority}'
    })


# Helper function
def get_vendor_from_order(order):
    """Get vendor from order - you'll need to implement this based on your Order model"""
    try:
        for item in order.items.filter(product__isnull=False):
            if (item.product.creator and
                    hasattr(item.product.creator, 'userprofile') and
                    hasattr(item.product.creator.userprofile, 'Coach_profile')):
                return item.product.creator
        return None
    except Exception:
        return None